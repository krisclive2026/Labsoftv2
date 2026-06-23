from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
import sqlite3
import os
import json
import pdfplumber
from datetime import datetime, timedelta
import uuid
import hashlib
import base64
import platform
from cryptography.hazmat.primitives.asymmetric import rsa, padding
from cryptography.hazmat.primitives import hashes, serialization
from cryptography.hazmat.backends import default_backend
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import Image
from reportlab.pdfgen import canvas
import tempfile
import io
from xhtml2pdf import pisa
 
import os
import sys
from flask import Flask, render_template
from flask_cors import CORS
 
# --------------------------
# Detect base directory
# --------------------------
if getattr(sys, 'frozen', False):
    # Running as PyInstaller EXE — bundled assets live in _MEIPASS
    BUNDLE_DIR = sys._MEIPASS
    EXE_DIR    = os.path.dirname(sys.executable)
else:
    # Running as plain Python script
    BUNDLE_DIR = os.path.dirname(os.path.abspath(__file__))
    EXE_DIR    = BUNDLE_DIR

# --------------------------
# Read-only bundled assets (inside the EXE / source tree)
# templates/ and static/ are embedded via --add-data
# --------------------------
TEMPLATE_DIR = os.path.join(BUNDLE_DIR, "templates")
STATIC_DIR   = os.path.join(BUNDLE_DIR, "static")

# --------------------------
# Writable data directory
# - Script mode : same folder as app.py (existing layout, no migration needed)
# - EXE mode    : hidden %APPDATA%\.labsoft\ so data survives EXE updates
# --------------------------
if getattr(sys, 'frozen', False):
    # Running as PyInstaller EXE
    if os.name == 'nt':
        _APPDATA = os.environ.get('APPDATA', os.path.expanduser('~'))
        DATA_DIR = os.path.join(_APPDATA, '.labsoft')
    else:
        DATA_DIR = os.path.join(os.path.expanduser('~'), '.labsoft')
else:
    # Running as plain script — keep everything next to app.py
    DATA_DIR = EXE_DIR

DATABASE_DIR  = os.path.join(DATA_DIR, "databases")
UPLOAD_FOLDER = os.path.join(DATA_DIR, "uploads")
LOGO_FOLDER   = os.path.join(DATA_DIR, "logos")
DB_PATH          = os.path.join(DATABASE_DIR, "lab_categories_fixed.db")
PATIENTS_DB_PATH = os.path.join(DATABASE_DIR, "patients.db")
DOCTORS_DB_PATH  = os.path.join(DATABASE_DIR, "doctors.db")

# --------------------------
# Create required writable folders
# --------------------------
os.makedirs(DATABASE_DIR, exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(LOGO_FOLDER, exist_ok=True)

# On Windows: mark .labsoft as a hidden system folder
if os.name == 'nt':
    try:
        import ctypes
        ctypes.windll.kernel32.SetFileAttributesW(DATA_DIR, 0x02)  # FILE_ATTRIBUTE_HIDDEN
    except Exception:
        pass

# --------------------------
# Migrate old license/data files from EXE_DIR to DATA_DIR
# (handles users upgrading from the old path layout)
# --------------------------
def _migrate_old_files():
    import shutil
    _migrate_list = [
        ("license.dat", DATA_DIR),
        ("license.clk", DATA_DIR),
        ("license.ses", DATA_DIR),
        ("databases",   DATA_DIR),   # entire folder
    ]
    for name, dest in _migrate_list:
        src = os.path.join(EXE_DIR, name)
        dst = os.path.join(dest, name)
        if os.path.exists(src) and not os.path.exists(dst):
            try:
                if os.path.isdir(src):
                    shutil.copytree(src, dst)
                else:
                    shutil.copy2(src, dst)
            except Exception:
                pass  # non-fatal — fresh activation will be needed

_migrate_old_files()
 
# --------------------------
# Initialize Flask app
# --------------------------
app = Flask(__name__, template_folder=TEMPLATE_DIR, static_folder=STATIC_DIR)
CORS(app)
 
# ── LICENSE ENFORCEMENT — block all API routes when license is invalid ─────────
UNPROTECTED_ROUTES = {
    'index', 'license_status', 'license_activate', 'license_machine_id',
    'get_session_count', 'static'
}
 
@app.before_request
def enforce_license():
    """Block every API route unless the license is valid and not expired."""
    endpoint = request.endpoint or ''
    # Allow unprotected routes through
    if endpoint in UNPROTECTED_ROUTES:
        return None
    # Only enforce on /api/* paths
    if not request.path.startswith('/api/'):
        return None
    # DISABLED FOR DEVELOPMENT - Remove this before production release
    # status = check_license_status()
    # if not status.get('valid'):
    #     reason = status.get('message', 'License is not valid.')
    #     return jsonify({'error': 'license_invalid', 'message': reason}), 403
    return None
 
# ══════════════════════════════════════════════════════════════════════════════
# LICENSE SYSTEM — RSA-based machine-locked licensing
# ══════════════════════════════════════════════════════════════════════════════
 
# RSA Public Key (embedded in app — private key stays with developer only)
RSA_PUBLIC_KEY_PEM = b"""-----BEGIN PUBLIC KEY-----
MIIBIjANBgkqhkiG9w0BAQEFAAOCAQ8AMIIBCgKCAQEA14UifELm2Xaw1XMXR31z
DL9TlKM4FIt1CqfUH2zaxb2GxctWlv1QOFS9v+GAIKv6RvQ3EuRb5UcD2Qqi0pTu
ILJU2bEWiwbGegvkjkH3rQu3YXBiY4yVOASRS8tgbDMRAUE5F36BmBXXNElzRdvK
/fNe/jXcXanht9VlDPqpiZo5pYD557xfdJoxR2spNIUjjU0Z+rXabj5WbmxLb4ZY
BhRxvOgNEjNTEbRd/J0Y4wAy/KUBMQgp+YbiKg6DHZF73e0AXsskHu7czsskR2LC
dD0BSXWGiepVwK2p+XZjLX3DztIZ8AFKX43ULXE/ZKemyEErUoYyVsxCj5DlatdH
GQIDAQAB
-----END PUBLIC KEY-----"""
 
# Shared secret for System Code — deterministic AES-ECB + base32
# Same key must be in license_keygen.py
_SC_KEY = bytes.fromhex("477a4b74394b346b6c613277787963320000000000000000000000000000001f")
 
def get_system_code(machine_id: str) -> str:
    """Encrypt machine ID into a deterministic, copy-safe System Code."""
    from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
    from cryptography.hazmat.backends import default_backend
    import base64 as _b64
    data = machine_id.encode().ljust(32, b'\x00')
    c = Cipher(algorithms.AES(_SC_KEY), modes.ECB(), backend=default_backend())
    enc = c.encryptor()
    ct = enc.update(data) + enc.finalize()
    return _b64.b32encode(ct).decode().rstrip('=')
 
def decode_system_code(system_code: str) -> str:
    """Decrypt System Code back to machine ID."""
    from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
    from cryptography.hazmat.backends import default_backend
    import base64 as _b64
    padded = system_code.strip().upper() + '=' * (-len(system_code) % 8)
    ct = _b64.b32decode(padded)
    c = Cipher(algorithms.AES(_SC_KEY), modes.ECB(), backend=default_backend())
    dec = c.decryptor()
    pt = dec.update(ct) + dec.finalize()
    return pt.rstrip(b'\x00').decode()
 
LICENSE_FILE    = os.path.join(DATA_DIR, "license.dat")
CLOCK_LEDGER    = os.path.join(DATA_DIR, "license.clk")   # hidden monotone timestamp ledger
SESSION_COUNTER = os.path.join(DATA_DIR, "license.ses")   # successful session counter
 
# ── Local time ────────────────────────────────────────────────────────────────
import struct
import time as _time
 
def _get_now():
    """Returns current UTC datetime from the local system clock."""
    return datetime.utcnow()
 
 
# ── Clock Ledger: monotone "last seen" timestamp ──────────────────────────────
# Sentinel returned when the ledger file exists but fails integrity check
_LEDGER_TAMPERED = object()
 
def _read_ledger():
    """Return (timestamp: float, key_fingerprint: str|None) on success.
    Returns (_LEDGER_TAMPERED, None) if file exists but HMAC fails (tampered/corrupt).
    Returns (0.0, None) only if the file genuinely does not exist yet.
    """
    if not os.path.exists(CLOCK_LEDGER):
        return 0.0, None          # never written — fresh install
    try:
        with open(CLOCK_LEDGER, 'rb') as f:
            raw = f.read()
        # Format: HMAC-SHA256(secret, payload) + ts_bytes(8) + fp_bytes(32)
        if len(raw) < 40:
            return _LEDGER_TAMPERED, None   # too short — corrupted
        stored_hmac = raw[:32]
        payload     = raw[32:]
        import hmac as _hmac
        secret = _ledger_secret()
        expected = _hmac.new(secret, payload, hashlib.sha256).digest()
        if not _hmac.compare_digest(stored_hmac, expected):
            return _LEDGER_TAMPERED, None   # HMAC mismatch — tampered
        ts = struct.unpack('d', payload[:8])[0]
        fp = payload[8:40].hex() if len(payload) >= 40 else None
        return ts, fp
    except Exception:
        return _LEDGER_TAMPERED, None       # unreadable — treat as tampered
 
def _write_ledger(ts: float, key_fingerprint: str = None):
    """Persist timestamp + key fingerprint to tamper-evident ledger."""
    import hmac as _hmac
    try:
        secret   = _ledger_secret()
        ts_bytes = struct.pack('d', ts)
        fp_bytes = bytes.fromhex(key_fingerprint) if key_fingerprint else (b'\x00' * 32)
        payload  = ts_bytes + fp_bytes
        mac      = _hmac.new(secret, payload, hashlib.sha256).digest()
        with open(CLOCK_LEDGER, 'wb') as f:
            f.write(mac + payload)
    except Exception:
        pass
 
def _ledger_secret():
    """Derive an HMAC secret from the machine ID so ledger is machine-specific."""
    mid = get_machine_id()
    return hashlib.sha256(f"LCK:{mid}:labsoft".encode()).digest()
 
# ── Session counter ───────────────────────────────────────────────────────────
def _session_secret():
    mid = get_machine_id()
    return hashlib.sha256(f"SES:{mid}:labsoft".encode()).digest()
 
def _read_session_count() -> int:
    """Return the stored successful session count, or 0 if not yet created."""
    import hmac as _hmac
    try:
        if not os.path.exists(SESSION_COUNTER):
            return 0
        with open(SESSION_COUNTER, 'rb') as f:
            raw = f.read()
        # Format: HMAC-SHA256(secret, count_bytes) + count_bytes (8 bytes uint64)
        if len(raw) < 40:
            return 0
        stored_mac = raw[:32]
        count_bytes = raw[32:40]
        secret = _session_secret()
        expected = _hmac.new(secret, count_bytes, hashlib.sha256).digest()
        if not _hmac.compare_digest(stored_mac, expected):
            return 0    # tampered — reset silently
        return struct.unpack('Q', count_bytes)[0]
    except Exception:
        return 0
 
def _write_session_count(count: int):
    """Persist session count to HMAC-protected file."""
    import hmac as _hmac
    try:
        secret = _session_secret()
        count_bytes = struct.pack('Q', count)
        mac = _hmac.new(secret, count_bytes, hashlib.sha256).digest()
        with open(SESSION_COUNTER, 'wb') as f:
            f.write(mac + count_bytes)
    except Exception:
        pass
 
def increment_session_count() -> int:
    """Increment and persist the session counter. Returns the new count."""
    count = _read_session_count() + 1
    _write_session_count(count)
    return count
 
def _check_clock_rollback(now_dt: datetime, current_fp: str = None) -> str | None:
    """
    Compare now_dt against the ledger.
    Returns an error string if rollback or tamper detected, else None.
    Updates ledger if time moved forward (preserving stored fingerprint).
    """
    now_ts   = now_dt.timestamp()
    last_ts, stored_fp = _read_ledger()
    TOLERANCE = 86400  # allow 1 day tolerance (DST, time-zone changes)
 
    # Ledger exists but is tampered/corrupted
    # If we got here, license.dat exists and is valid — reset the corrupt ledger
    # and rewrite it with the current time so the user isn't hard-blocked.
    if last_ts is _LEDGER_TAMPERED:
        try:
            os.remove(CLOCK_LEDGER)
        except Exception:
            pass
        _write_ledger(now_dt.timestamp(), current_fp)
        return None  # treated as clean slate
 
    if last_ts > 0 and (now_ts + TOLERANCE) < last_ts:
        delta_days = int((last_ts - now_ts) / 86400)
        return (f"Clock rollback detected: system time is ~{delta_days} day(s) "
                f"behind the last recorded time. "
                f"Please restore the correct system date and time.")
 
    # Advance ledger; keep existing fingerprint unless a new one is provided
    fp_to_store = current_fp if current_fp is not None else stored_fp
    if now_ts > last_ts:
        _write_ledger(now_ts, fp_to_store)
    return None
 
# ── Machine ID ────────────────────────────────────────────────────────────────
def get_machine_id():
    """Generate a stable machine fingerprint from MAC address + hostname."""
    mac = uuid.getnode()
    hostname = platform.node()
    raw = f"{mac}:{hostname}"
    fingerprint = hashlib.sha256(raw.encode()).hexdigest()[:16].upper()
    return '-'.join([fingerprint[i:i+4] for i in range(0, 16, 4)])
 
# ── Core license verification ─────────────────────────────────────────────────
def _decode_license_key(license_key_b64):
    """Decode self-contained key → (expiry_str, nonce_bytes, signature_bytes).
    Supports both old format (10+256) and new format with nonce (10+8+256).
    """
    raw = base64.b64decode(license_key_b64)
    expiry_str = raw[:10].decode('ascii')   # e.g. "2026-04-11"
    rest = raw[10:]
    if len(rest) == 256:
        # Legacy format (no nonce)
        return expiry_str, b'', rest
    elif len(rest) == 264:
        # New format: 8-byte nonce + 256-byte signature
        return expiry_str, rest[:8], rest[8:]
    else:
        raise ValueError(f"Invalid key length: {len(raw)} bytes")
 
def _rsa_verify(machine_id, license_key_b64, expiry=None):
    """RSA signature check. Expiry and nonce are decoded from key itself."""
    public_key = serialization.load_pem_public_key(RSA_PUBLIC_KEY_PEM, backend=default_backend())
    expiry_from_key, nonce, sig = _decode_license_key(license_key_b64)
    if nonce:
        payload = f"{machine_id}:{expiry_from_key}:".encode() + nonce
    else:
        payload = f"{machine_id}:{expiry_from_key}".encode()   # legacy
    public_key.verify(sig, payload, padding.PKCS1v15(), hashes.SHA256())
 
def verify_license_key(machine_id, license_key_b64):
    """
    Full verification including clock-tamper detection.
    Returns (valid: bool, expiry: str|None, message: str)
    """
    try:
        if not os.path.exists(LICENSE_FILE):
            return False, None, "No license file found."
 
        with open(LICENSE_FILE, 'r') as f:
            data = json.load(f)
 
        stored_machine = data.get('machine_id', '')
        license_key    = data.get('license_key', license_key_b64)
 
        if stored_machine != machine_id:
            return False, None, "License is not valid for this machine."
 
        # 1) RSA signature check + decode expiry from key
        _rsa_verify(machine_id, license_key)
        expiry, _, __ = _decode_license_key(license_key)
 
        # 2) Get trusted time
        now_dt = _get_now()
 
        # 3) Clock rollback check against ledger
        rollback_err = _check_clock_rollback(now_dt)
        if rollback_err:
            return False, expiry, f"⚠️ {rollback_err}"
 
        # 4) Expiry check using trusted time
        expiry_dt = datetime.strptime(expiry, "%Y-%m-%d")
        if now_dt.date() > expiry_dt.date():
            return False, expiry, "License has expired. Please request a renewal key."
 
        # 5) Update last-seen in license file (monotone record)
        data['last_check'] = now_dt.strftime("%Y-%m-%d %H:%M:%S")
        with open(LICENSE_FILE, 'w') as f:
            json.dump(data, f)
 
        return True, expiry, "License is valid."
 
    except Exception as e:
        return False, None, f"License verification failed: {str(e)}"
 
def _key_fingerprint(license_key_b64: str) -> str:
    """SHA-256 hex fingerprint of a license key (used for one-time-use enforcement)."""
    return hashlib.sha256(license_key_b64.encode()).hexdigest()
 
def activate_license(machine_id, license_key_b64, expiry=None):
    """Activate and store license after verifying RSA signature and clock."""
    try:
        # RSA check + extract expiry from key itself
        _rsa_verify(machine_id, license_key_b64)
        expiry, _, __ = _decode_license_key(license_key_b64)
 
        # Compute fingerprint of this key
        new_fp = _key_fingerprint(license_key_b64)
 
        # ── One-time-use check ──────────────────────────────────────────────
        # 1. Check ledger — survives license.dat deletion
        _, ledger_fp = _read_ledger()
        if ledger_fp and ledger_fp == new_fp:
            return False, "This license key has already been used. Please request a new key."
 
        # 2. Check license.dat — catches replay within same install
        if os.path.exists(LICENSE_FILE):
            try:
                with open(LICENSE_FILE, 'r') as f:
                    existing = json.load(f)
                if existing.get('key_fingerprint') == new_fp:
                    return False, "This license key has already been used. Please request a new key."
            except Exception:
                pass
        # ───────────────────────────────────────────────────────────────────
 
        # Trusted time
        now_dt = _get_now()
 
        # Rollback check
        rollback_err = _check_clock_rollback(now_dt, new_fp)
        if rollback_err:
            return False, f"⚠️ {rollback_err}"
 
        expiry_dt = datetime.strptime(expiry, "%Y-%m-%d")
        if now_dt.date() > expiry_dt.date():
            return False, "The license key has already expired."
 
        license_data = {
            "machine_id":     machine_id,
            "license_key":    license_key_b64,
            "key_fingerprint": new_fp,
            "expiry":         expiry,
            "activated_at":   now_dt.strftime("%Y-%m-%d %H:%M:%S"),
            "last_check":     now_dt.strftime("%Y-%m-%d %H:%M:%S"),
        }
        with open(LICENSE_FILE, 'w') as f:
            json.dump(license_data, f)
 
        # Persist fingerprint + activation time to tamper-evident ledger
        _write_ledger(now_dt.timestamp(), new_fp)
        return True, "License activated successfully!"
    except Exception:
        return False, "Activation failed: Invalid key or wrong machine."
 
def check_license_status():
    """Returns dict: { valid, machine_id, expiry, days_left, message, clock_ok }"""
    mid = get_machine_id()
 
    # ── Tamper detection: check ledger first ─────────────────────────────────
    ledger_ts, ledger_fp = _read_ledger()
    ledger_exists  = os.path.exists(CLOCK_LEDGER)
    license_exists = os.path.exists(LICENSE_FILE)
 
    # Ledger is present but corrupted/tampered
    if ledger_ts is _LEDGER_TAMPERED:
        # If license.dat is also missing, this is likely a stale/corrupt file
        # from a previous install — delete it and treat as a fresh install.
        if not license_exists:
            try:
                os.remove(CLOCK_LEDGER)
            except Exception:
                pass
            return {"valid": False, "machine_id": mid, "expiry": None, "days_left": 0,
                    "message": "No license found. Please activate.",
                    "expired": False, "clock_ok": True}
        # Ledger tampered AND license.dat present — genuine integrity violation
        return {"valid": False, "machine_id": mid, "expiry": None, "days_left": 0,
                "message": "⚠️ License integrity check failed. Please contact support to reactivate.",
                "expired": False, "clock_ok": False}
 
    # Ledger has a fingerprint (activation happened) but license.dat is gone
    if ledger_fp and not license_exists:
        return {"valid": False, "machine_id": mid, "expiry": None, "days_left": 0,
                "message": "⚠️ License file has been removed. Please contact support to reactivate.",
                "expired": False, "clock_ok": False}
 
    # Ledger has a fingerprint but license.dat fingerprint doesn't match
    if ledger_fp and license_exists:
        try:
            with open(LICENSE_FILE, 'r') as f:
                _dat = json.load(f)
            if _dat.get('key_fingerprint') != ledger_fp:
                return {"valid": False, "machine_id": mid, "expiry": None, "days_left": 0,
                        "message": "⚠️ License file has been modified. Please contact support to reactivate.",
                        "expired": False, "clock_ok": False}
        except Exception:
            return {"valid": False, "machine_id": mid, "expiry": None, "days_left": 0,
                    "message": "⚠️ License file is corrupted. Please contact support to reactivate.",
                    "expired": False, "clock_ok": False}
 
    # license.dat present but ledger missing → ledger was deleted
    if license_exists and not ledger_exists:
        return {"valid": False, "machine_id": mid, "expiry": None, "days_left": 0,
                "message": "⚠️ License clock file has been removed. Please contact support to reactivate.",
                "expired": False, "clock_ok": False}
 
    # No ledger, no license → genuine fresh install, show activation
    if not ledger_exists and not license_exists:
        return {"valid": False, "machine_id": mid, "expiry": None, "days_left": 0,
                "message": "No license found. Please activate.", "expired": False,
                "clock_ok": True}
    # ─────────────────────────────────────────────────────────────────────────
 
    try:
        with open(LICENSE_FILE, 'r') as f:
            data = json.load(f)
 
        valid, expiry, msg = verify_license_key(mid, data.get('license_key', ''))
        days_left = 0
        expired   = False
        clock_ok  = "rollback" not in msg.lower() and "clock" not in msg.lower() and "integrity" not in msg.lower()
 
        if expiry:
            try:
                now_dt = _get_now()
            except Exception:
                now_dt = datetime.utcnow()
            try:
                expiry_dt  = datetime.strptime(expiry, "%Y-%m-%d")
                days_left  = max(0, (expiry_dt - now_dt).days)
                expired    = now_dt.date() > expiry_dt.date()
            except Exception:
                pass
 
        return {"valid": valid, "machine_id": mid, "expiry": expiry,
                "days_left": days_left, "message": msg, "expired": expired,
                "clock_ok": clock_ok}
    except Exception as e:
        return {"valid": False, "machine_id": mid, "expiry": None, "days_left": 0,
                "message": f"License check error: {str(e)}", "expired": False,
                "clock_ok": True}
 
# ── LICENSE API ROUTES ────────────────────────────────────────────────────────
 
# --------------------------
# Debug prints (optional)
# --------------------------
print("Bundle Directory:", BUNDLE_DIR)
print("Data Directory:", DATA_DIR)
print("Template Folder Exists:", os.path.exists(TEMPLATE_DIR))
print("Static Folder Exists:", os.path.exists(STATIC_DIR))
print("Database Path:", DB_PATH)
print("Upload Folder:", UPLOAD_FOLDER)
print("Logo Folder:", LOGO_FOLDER)
 
# ── LOGO CANVASMAKER HELPERS ──────────────────────────────────────────────────
def get_logo_path():
    """Get active logo path from settings (with fallback)."""
    conn = get_db()
    row = conn.execute("SELECT logo_path FROM lab_settings LIMIT 1").fetchone()
    conn.close()
    return row['logo_path'] if row and row['logo_path'] and os.path.exists(row['logo_path']) else None
 
def add_lab_logo_canvas(doc):
    """Canvasmaker: Draw lab logo (top-right) + end_of_report.png (footer)."""
    logo_path = get_logo_path()
    end_logo = os.path.join(LOGO_FOLDER, "end_of_report.png")
    
    def _draw(canvas, doc):
        w, h = doc.pagesize
        
        # Lab logo (top-right, if exists)
        if logo_path:
            try:
                img_w, img_h = 40, 25
                x = w - 15*mm - img_w
                y = h - 15*mm - img_h
                canvas.drawImage(logo_path, x, y, width=img_w, height=img_h, mask='auto')
            except Exception:
                pass  # silently skip missing/broken logo
        
        # End-of-report logo (footer center)
        if os.path.exists(end_logo):
            try:
                img_w, img_h = 60, 25
                x = (w - img_w) / 2
                y = 15*mm
                canvas.drawImage(end_logo, x, y, width=img_w, height=img_h, mask='auto')
            except Exception:
                pass
    
    return _draw
# ── DB INIT ──────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn
 
def get_db_joined():
    """Main DB with patients.db and doctors.db ATTACHed so cross-DB JOINs work."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute(f"ATTACH DATABASE '{PATIENTS_DB_PATH}' AS pdb")
    conn.execute(f"ATTACH DATABASE '{DOCTORS_DB_PATH}'  AS ddb")
    # Create views so existing SQL (patients p / doctors d) keeps working
    conn.execute("""
        CREATE TEMP VIEW IF NOT EXISTS patients AS
        SELECT * FROM pdb.patients
    """)
    conn.execute("""
        CREATE TEMP VIEW IF NOT EXISTS doctors AS
        SELECT * FROM ddb.doctors
    """)
    return conn
 
def get_patients_db():
    conn = sqlite3.connect(PATIENTS_DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn
 
def get_doctors_db():
    conn = sqlite3.connect(DOCTORS_DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn
 
def init_patients_db():
    """Create patients table in its own DB and migrate data from main DB if needed."""
    conn = get_patients_db()
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS patients (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        age INTEGER,
        gender TEXT,
        phone TEXT,
        email TEXT,
        address TEXT,
        blood_group TEXT,
        greeting TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    )''')
    conn.commit()
    # Add greeting column if migrating from an older DB that doesn't have it yet
    try:
        c.execute("ALTER TABLE patients ADD COLUMN greeting TEXT")
        conn.commit()
    except Exception:
        pass
    # Migrate from main DB if patients.db is empty
    c.execute("SELECT COUNT(*) FROM patients")
    if c.fetchone()[0] == 0 and os.path.exists(DB_PATH):
        try:
            main = sqlite3.connect(DB_PATH)
            rows = main.execute("SELECT id,name,age,gender,phone,email,address,blood_group,created_at FROM patients").fetchall()
            if rows:
                c.executemany("INSERT OR IGNORE INTO patients (id,name,age,gender,phone,email,address,blood_group,created_at) VALUES (?,?,?,?,?,?,?,?,?)", rows)
                conn.commit()
            main.close()
        except Exception:
            pass
    conn.close()
 
def init_doctors_db():
    """Create doctors table in its own DB and migrate data from main DB if needed."""
    conn = get_doctors_db()
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS doctors (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        specialization TEXT,
        hospital TEXT,
        phone TEXT,
        email TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    )''')
    conn.commit()
    # Migrate from main DB if doctors.db is empty
    c.execute("SELECT COUNT(*) FROM doctors")
    if c.fetchone()[0] == 0 and os.path.exists(DB_PATH):
        try:
            main = sqlite3.connect(DB_PATH)
            rows = main.execute("SELECT id,name,specialization,hospital,phone,email,created_at FROM doctors").fetchall()
            if rows:
                c.executemany("INSERT OR IGNORE INTO doctors (id,name,specialization,hospital,phone,email,created_at) VALUES (?,?,?,?,?,?,?)", rows)
                conn.commit()
            main.close()
        except Exception:
            pass
    conn.close()
 
def init_db():
    conn = get_db()
    c = conn.cursor()
    c.executescript('''
        CREATE TABLE IF NOT EXISTS patients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            age INTEGER,
            gender TEXT,
            phone TEXT,
            email TEXT,
            address TEXT,
            blood_group TEXT,
            greeting TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS doctors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            specialization TEXT,
            hospital TEXT,
            phone TEXT,
            email TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            patient_id INTEGER,
            doctor_id INTEGER,
            report_title TEXT,
            report_date TEXT,
            extracted_text TEXT,
            ai_interpretation TEXT,
            pdf_path TEXT,
            original_pdf_path TEXT,
            html_content TEXT,
            status TEXT DEFAULT 'pending',
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(patient_id) REFERENCES patients(id),
            FOREIGN KEY(doctor_id) REFERENCES doctors(id)
        );
        CREATE TABLE IF NOT EXISTS lab_settings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lab_name TEXT,
            lab_address TEXT,
            lab_phone TEXT,
            lab_email TEXT,
            logo_path TEXT,
            form_design TEXT
        );
        CREATE TABLE IF NOT EXISTS appointments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            patient_id INTEGER,
            doctor_id INTEGER,
            appointment_date TEXT,
            appointment_time TEXT,
            test_names TEXT,
            notes TEXT,
            status TEXT DEFAULT 'scheduled',
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(patient_id) REFERENCES patients(id),
            FOREIGN KEY(doctor_id) REFERENCES doctors(id)
        );
        CREATE TABLE IF NOT EXISTS test_catalog (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            test_name TEXT NOT NULL,
            category TEXT,
            unit TEXT,
            normal_min REAL,
            normal_max REAL,
            normal_text TEXT,
            description TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS test_stage1 (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS test_stage2 (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            stage1_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(stage1_id) REFERENCES test_stage1(id)
        );
        CREATE TABLE IF NOT EXISTS test_stage3 (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            stage2_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(stage2_id) REFERENCES test_stage2(id)
        );
        CREATE TABLE IF NOT EXISTS bills (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            patient_id INTEGER NOT NULL,
            doctor_id INTEGER,
            bill_date TEXT,
            bill_type TEXT DEFAULT 'Cash',
            cheque_ref TEXT,
            total_amount REAL DEFAULT 0,
            status TEXT DEFAULT 'saved',
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(patient_id) REFERENCES patients(id),
            FOREIGN KEY(doctor_id) REFERENCES doctors(id)
        );
        CREATE TABLE IF NOT EXISTS bill_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bill_id INTEGER NOT NULL,
            stage1_name TEXT,
            stage2_name TEXT,
            stage3_name TEXT,
            test_name TEXT,
            rate REAL DEFAULT 0,
            FOREIGN KEY(bill_id) REFERENCES bills(id)
        );
        CREATE TABLE IF NOT EXISTS test_rates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            stage1_name TEXT,
            stage2_name TEXT,
            stage3_name TEXT,
            rate REAL DEFAULT 0,
            UNIQUE(stage1_name, stage2_name, stage3_name)
        );
        CREATE TABLE IF NOT EXISTS other_labs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS other_lab_ranges (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lab_id INTEGER NOT NULL,
            test_name TEXT NOT NULL,
            category TEXT,
            unit TEXT,
            normal_min_m REAL,
            normal_max_m REAL,
            normal_text_m TEXT,
            normal_min_f REAL,
            normal_max_f REAL,
            normal_text_f TEXT,
            normal_min_c REAL,
            normal_max_c REAL,
            normal_text_c TEXT,
            description TEXT,
            interpretation TEXT,
            amount REAL DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(lab_id) REFERENCES other_labs(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS stage_exclusions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            level INTEGER NOT NULL,
            category TEXT NOT NULL DEFAULT '',
            sub_category TEXT NOT NULL DEFAULT '',
            test_name TEXT NOT NULL DEFAULT '',
            created_at TEXT DEFAULT (datetime('now')),
            UNIQUE(level, category, sub_category, test_name)
        );
    ''')
    # Insert default settings if none exist
    c.execute("SELECT COUNT(*) FROM lab_settings")
    if c.fetchone()[0] == 0:
        c.execute("INSERT INTO lab_settings (lab_name, lab_address, lab_phone, lab_email) VALUES (?,?,?,?)",
                  ("My Diagnostic Lab", "123 Health Street, City", "+91-9999999999", "lab@example.com"))
    conn.commit()
    conn.close()
 
def seed_test_catalog():
    conn = get_db()
    c = conn.cursor()
    # Add form_design column if not exists
    try:
        c.execute("ALTER TABLE lab_settings ADD COLUMN form_design TEXT")
        conn.commit()
    except Exception:
        pass
    # Add gender-specific range columns if they don't exist yet
    try:
        c.execute("ALTER TABLE test_catalog ADD COLUMN normal_min_m REAL")
        c.execute("ALTER TABLE test_catalog ADD COLUMN normal_max_m REAL")
        c.execute("ALTER TABLE test_catalog ADD COLUMN normal_min_f REAL")
        c.execute("ALTER TABLE test_catalog ADD COLUMN normal_max_f REAL")
        c.execute("ALTER TABLE test_catalog ADD COLUMN normal_min_c REAL")
        c.execute("ALTER TABLE test_catalog ADD COLUMN normal_max_c REAL")
        c.execute("ALTER TABLE test_catalog ADD COLUMN normal_text_m TEXT")
        c.execute("ALTER TABLE test_catalog ADD COLUMN normal_text_f TEXT")
        c.execute("ALTER TABLE test_catalog ADD COLUMN normal_text_c TEXT")
        c.execute("ALTER TABLE test_catalog ADD COLUMN sub_category TEXT")
        c.execute("ALTER TABLE test_catalog ADD COLUMN amount REAL DEFAULT 0")
        c.execute("ALTER TABLE test_catalog ADD COLUMN interpretation TEXT")
        conn.commit()
    except Exception:
        pass  # columns already exist
 
    c.execute("SELECT COUNT(*) FROM test_catalog")
    if c.fetchone()[0] == 0:
        # Full New Care Laboratory test catalog
        # Fields: test_name, category, unit, normal_min, normal_max, normal_text,
        #         description, normal_min_m, normal_max_m, normal_min_f, normal_max_f,
        #         normal_min_c, normal_max_c, normal_text_m, normal_text_f, normal_text_c,
        #         amount, interpretation
        tests = [
            # ── BIO-CHEMISTRY ─────────────────────────────────────────────────────
            ('RANDOM BLOOD SUGAR', 'BIO-CHEMISTRY', 'mg/dl', 80, 130, None,
             'Random blood glucose level',
             80, 130, 80, 130, None, None, '80 - 130', '80 - 130', None, 30,
             'Fasting Blood Sugar more than 126 mg/dl on more than one occasion can indicate Diabetes Mellitus.'),
 
            ('FASTING BLOOD SUGAR', 'BIO-CHEMISTRY', 'mg/dl', 70, 110, None,
             'Fasting blood glucose level',
             70, 110, 70, 110, None, None, '70 - 110', '70 - 110', None, 30,
             'Fasting Blood Sugar more than 126 mg/dl on more than one occasion can indicate Diabetes Mellitus.'),
 
            ('POST PRANDIAL BLOOD SUGAR', 'BIO-CHEMISTRY', 'mg/dl', 90, 140, None,
             'Post-meal blood glucose level',
             90, 140, 90, 140, None, None, '90 - 140', '90 - 140', None, 30, None),
 
            ('HbA1c', 'BIO-CHEMISTRY', '%', 5.7, 6.4, 'Normal: 4 - 5.6%, Pre-diabetes: 5.7 - 6.4%, Diabetes: 6.5% or higher',
             'Glycated haemoglobin — 3-month average blood sugar',
             5.7, 6.4, 5.5, 6.4, 5.7, 6.4,
             'Normal: 4 - 5.6%\nPre-diabetes: 5.7 - 6.4%\nDiabetes: 6.5% or higher',
             'Normal: 4 - 5.6%\nPre-diabetes: 5.7 - 6.4%\nDiabetes: 6.5% or higher',
             'Normal: 4 - 5.6%\nPre-diabetes: 5.7 - 6.4%\nDiabetes: 6.5% or higher', 450,
             'HbA1C level reflects the mean glucose concentration over previous 8-12 weeks and provides better indication of long term glycemic control.'),
 
            ('GLYCOSYLATED HB (HBA1c)', 'BIO-CHEMISTRY', '%', 4, 6.5, 'Non Diabetic: 3.0-6.0%, Pre-diabetes: 5.7-6.4%, Diabetes: 6.5% or higher',
             'Glycosylated haemoglobin',
             4, 6.5, 4, 6.5, 4, 6.4, None, None, None, 500, None),
 
            ('UREA', 'BIO-CHEMISTRY', 'mg%', 10, 40, None,
             'Blood urea — kidney function marker',
             10, 40, 10, 40, 10, 40, '10 - 40', '10 - 40', '10 - 40', 100, None),
 
            ('CREATININE', 'BIO-CHEMISTRY', 'mg/dl', 0.6, 1.2, None,
             'Serum creatinine — kidney filtration marker',
             0.5, 1.4, 0.5, 1.2, 0.2, 1.2, '0.5 - 1.4', '0.5 - 1.2', '0.2 - 1.2', 100, None),
 
            ('URIC ACID', 'BIO-CHEMISTRY', 'mg/dl', 3.5, 7.2, None,
             'Uric acid — purine breakdown product',
             4.0, 7.2, 2.7, 6.5, None, None, '4.0 - 7.2', '2.7 - 6.5', 'URIC ACID', 100, None),
 
            ('BLOOD UREA NITROGEN', 'BIO-CHEMISTRY', 'mg/dl', 6, 20, None,
             'Blood urea nitrogen',
             6, 20, 6, 20, None, None, '6 - 20', '6 - 20', None, 250, None),
 
            ('RENAL FUNCTION TEST', 'BIO-CHEMISTRY', '', None, None, 'Urea + Creatinine + Uric Acid',
             'Complete kidney function panel',
             None, None, None, None, None, None, None, None, None, 0, None),
 
            ('SGOT', 'BIO-CHEMISTRY', 'IU/L', 8, 33, None,
             'Serum glutamic oxaloacetic transaminase (AST)',
             8, 33, 8, 33, None, None, '8 - 33', '8 - 33', None, 150, None),
 
            ('SGPT', 'BIO-CHEMISTRY', 'IU/L', 0, 38, None,
             'Serum glutamic pyruvic transaminase (ALT)',
             0, 38, 0, 38, None, None, '0 - 38', '0 - 38', None, 150, None),
 
            ('ALKALINE PHOSPHATASE', 'BIO-CHEMISTRY', 'IU/L', 70, 175, None,
             'Bone and liver enzyme',
             70, 175, 70, 175, None, None, '70 - 175', '70 - 175', None, 100, None),
 
            ('TOTAL BILIRUBIN', 'BIO-CHEMISTRY', 'mg/dl', 0.3, 1.1, None,
             'Total bilirubin — liver and red cell breakdown',
             0.3, 1.1, 0.3, 1.1, None, None, '0.3 - 1.1', '0.3 - 1.1', None, 0, None),
 
            ('BILIRUBIN', 'BIO-CHEMISTRY', 'mg/dl', None, None, 'Direct: 0.1-0.3, Indirect: 0.2-0.7',
             'Bilirubin panel',
             None, None, None, None, None, None, None, None, None, 0, None),
 
            ('TOTAL PROTEIN', 'BIO-CHEMISTRY', 'gm/dL', 5.5, 6.5, None,
             'Total serum protein',
             5.5, 6.5, 5.5, 6.5, None, None, '5.5 - 6.5', '5.5 - 6.5', None, 100, None),
 
            ('ALBUMIN', 'BIO-CHEMISTRY', 'mg/dl', 3.5, 5.2, None,
             'Serum albumin — main blood protein',
             3.5, 5.2, 3.5, 5.2, None, None, '3.5 - 5.2', '3.5 - 5.2', None, 100, None),
 
            ('LIVER FUNCTION TEST', 'BIO-CHEMISTRY', '', None, None, 'Comprehensive LFT panel',
             'Comprehensive liver function panel',
             None, None, None, None, None, None, None, None, None, 0, None),
 
            ('LIPID PROFILE', 'BIO-CHEMISTRY', '', None, None, 'Total cholesterol, LDL, HDL, TGL, VLDL',
             'Complete lipid panel',
             None, None, None, None, None, None, None, None, None, 0, None),
 
            ('TOTAL CHOLESTEROL', 'BIO-CHEMISTRY', 'mg/dl', None, 200, 'Desirable: <200, Borderline: 200-239, High: >=240',
             'Total blood cholesterol',
             200, 239, 200, 239, 200, 239,
             'Desirable Report: <200\nBorderline High: 200-239\nHigh: >=240',
             'Desirable Report: <200\nBorderline High: 200-239\nHigh: >=240',
             'Desirable Report: <200\nBorderline High: 200-239\nHigh: >=240', 125, None),
 
            ('LDL-CHOLESTEROL', 'BIO-CHEMISTRY', 'mg/dl', 70, 100, 'Optimal: <100, Near Optimal: 100-129, Borderline High: 130-159',
             'Low-density lipoprotein cholesterol',
             70, 100, 70, 100, None, None,
             'Optimal: <100\nNear Optimal: 100-129\nBorderline High: 130-159\nHigh: 160-189\nVery High: >=190',
             'Optimal: <100\nNear Optimal: 100-129\nBorderline High: 130-159\nHigh: 160-189\nVery High: >=190',
             None, 100, None),
 
            ('HDL-CHOLESTEROL', 'BIO-CHEMISTRY', 'mg/dl', 35, 55, None,
             'High-density lipoprotein cholesterol',
             35, 55, 35, 55, None, None, '35 - 55', '35 - 55', None, 100, None),
 
            ('TRIGLYCERIDES (TGL)', 'BIO-CHEMISTRY', 'mg/dl', 70, 175, 'Normal: <150, Borderline: 150-199, High: 200-499',
             'Blood triglycerides',
             70, 175, 70, 175, None, None,
             'Normal: <150\nBorderline High: 150-199\nHigh: 200-499\nVery High: >=500',
             'Normal: <150\nBorderline High: 150-199\nHigh: 200-499\nVery High: >=500',
             None, 100, None),
 
            ('VLDL-CHOLESTEROL', 'BIO-CHEMISTRY', 'mg/dl', None, 40, '< 40',
             'Very low-density lipoprotein',
             None, 40, None, 40, None, None, '< 40', '< 40', None, 100, None),
 
            ('CALCIUM', 'BIO-CHEMISTRY', 'mg/dl', 8.5, 10.2, None,
             'Serum calcium',
             8.5, 10.2, 8.5, 10.2, None, None, '8.5 - 10.2', '8.5 - 10.2', None, 300, None),
 
            ('MAGNESIUM', 'BIO-CHEMISTRY', 'mg/dl', 1.7, 2.2, None,
             'Serum magnesium',
             1.7, 2.2, 1.7, 2.2, None, None, '1.7 - 2.2', '1.7 - 2.2', None, 500, None),
 
            ('SODIUM', 'BIO-CHEMISTRY', 'mmol/L', 135, 148, None,
             'Serum sodium',
             135, 148, 135, 148, None, None, '135 - 148', '135 - 148', None, 150, None),
 
            ('POTASSIUM', 'BIO-CHEMISTRY', 'mmol/L', 3.5, 5.1, None,
             'Serum potassium',
             3.5, 5.1, 3.5, 5.1, None, None, '3.5 - 5.1', '3.5 - 5.1', None, 150, None),
 
            ('ELECTROLYTES, SERUM', 'BIO-CHEMISTRY', 'mmol/L', None, None, 'Na: 135-148, K: 3.5-5.1, Cl: 96-106',
             'Serum electrolytes panel',
             None, None, None, None, None, None, None, None, None, 0, None),
 
            ('C-REACTIVE PROTEIN (CRP)', 'BIO-CHEMISTRY', 'mg/L', 0, 6, None,
             'Inflammation marker',
             0, 6, 0, 6, 0, 56, '0 - 6', '0 - 6', '0 - 56', 500,
             'CRP levels can increase dramatically after severe trauma, bacterial infection, inflammation, surgery, or neoplastic proliferation. Elevated values are consistent with an acute inflammatory process.'),
 
            ('IRON', 'BIO-CHEMISTRY', 'mcg/dL', 60, 170, None,
             'Serum iron',
             60, 170, 60, 170, None, None, '60 - 170', '60 - 170', None, 950, None),
 
            ('AMYLASE', 'BIO-CHEMISTRY', 'U/L', 40, 140, None,
             'Pancreatic enzyme — elevated in pancreatitis',
             40, 140, 40, 140, None, None, '40 - 140', '40 - 140', None, 500, None),
 
            ('LIPASE', 'BIO-CHEMISTRY', 'U/L', 60, 140, None,
             'Pancreatic enzyme',
             60, 140, 60, 140, None, None, '60 - 140', '60 - 140', None, 500, None),
 
            ('D-DIMER', 'BIO-CHEMISTRY', '', 0, 0.50, None,
             'Fibrin degradation product — clotting marker',
             0, 0.50, 0, 0.50, None, None, '0 - 0.50', '0 - 0.50', None, 1500, None),
 
            ('TROPONIN I', 'BIO-CHEMISTRY', 'ng/mL', 0, 0.04, None,
             'Cardiac muscle damage marker',
             0, 0.04, 0, 0.04, None, None, '0 - 0.04', '0 - 0.04', None, 850, None),
 
            ('TROPONIN T', 'BIO-CHEMISTRY', 'ng/mL', 0, 0.04, None,
             'Cardiac muscle damage marker',
             0, 0.04, 0, 0.04, None, None, '0 - 0.04', '0 - 0.04', None, 850, None),
 
            ('VITAMIN B-12 LEVEL', 'BIO-CHEMISTRY', 'pg/mL', 187, 883, None,
             'Cobalamin — nerve and blood cell health',
             187, 883, 187, 883, None, None, '187 - 883', '187 - 883', None, 950,
             'Vitamin B12 deficiency can cause Megaloblastic anemia, nerve damage and degeneration of the spinal cord.'),
 
            ('VITAMIN D3 LEVEL', 'BIO-CHEMISTRY', 'ng/mL', 20, 50, None,
             '25-hydroxy vitamin D',
             20, 50, 20, 50, None, None, '20 - 50', '20 - 50', None, 1500, None),
 
            ('PROTHROMBIN TIME (PT)', 'BIO-CHEMISTRY', 'Seconds', 11, 17, None,
             'Clotting time test',
             11, 17, 11, 17, None, None, '11 - 17', '11 - 17', None, 0, None),
 
            ('PROTEIN CREATININE RATIO', 'BIO-CHEMISTRY', '', None, None,
             'Less Than 0.20: NORMAL\n0.20-1.00: LOW GRADE PROTEINURIA\n1.01-5.00: MODERATE PROTEINURIA\nMore Than 5.01: NEPHROSIS',
             'Urine protein:creatinine ratio',
             None, None, None, None, None, None, None, None, None, 0, None),
 
            ('URINE MICROALBUMIN', 'BIO-CHEMISTRY', 'mg/dl', 20, 30, None,
             'Urine microalbumin — early kidney damage marker',
             20, 30, 20, 30, None, None, '20 - 30', '20 - 30', None, 550, None),
 
            # ── HAEMATOLOGY ────────────────────────────────────────────────────────
            ('COMPLETE BLOOD COUNT (CBC)', 'HAEMATOLOGY', '', None, None, 'CBC with differential',
             'Full blood count panel',
             None, None, None, None, None, None, None, None, None, 0,
             'It indicates presence and intensity of an inflammatory process. Used to monitor the course or response to treatment of diseases like tuberculosis, acute rheumatic fever. It is also increased in multiple myeloma, hypothyroidism.'),
 
            ('HAEMOGLOBIN', 'HAEMATOLOGY', 'gms%', 12, 17, None,
             'Oxygen-carrying protein in red blood cells',
             13, 17, 12, 14.5, 4, 12, '13 - 17', '12 - 14.5', '4 - 12', 0, None),
 
            ('RED BLOOD CELLS (RBC)', 'HAEMATOLOGY', 'millions/cumm', 4.3, 5.8, None,
             'Red blood cell count',
             4.5, 5.5, 4.3, 5.8, 4, 5.2, '4.5 - 5.5', '4.3 - 5.8', '4 - 5.2', 0, None),
 
            ('TOTAL WBC COUNT', 'HAEMATOLOGY', 'Cells/cumm', 4000, 11000, None,
             'Total white blood cell count',
             4000, 11000, 4000, 11000, 4000, 11000, '4000 - 11000', '4000 - 11000', '4000 - 11000', 0, None),
 
            ('NEUTROPHIL', 'HAEMATOLOGY', '%', 40, 70, None,
             'Neutrophil differential count',
             40, 70, 40, 70, None, None, '40 - 70', '40 - 70', None, 0, None),
 
            ('LYMPHOCYTES', 'HAEMATOLOGY', '%', 20, 40, None,
             'Lymphocyte differential count',
             20, 45, 20, 45, None, None, '20 - 45', '20 - 45', None, 0, None),
 
            ('EOSINOPHIL', 'HAEMATOLOGY', '%', 0, 8, None,
             'Eosinophil differential count',
             2, 10, 2, 10, None, None, '2 - 10', '2 - 10', None, 0, None),
 
            ('MONOCYTE', 'HAEMATOLOGY', '%', 0, 2, None,
             'Monocyte differential count',
             1, 6, 1, 6, None, None, '1 - 6', '1 - 6', None, 0, None),
 
            ('BASOPHIL', 'HAEMATOLOGY', '%', 0, 1, None,
             'Basophil differential count',
             0, 1, 0, 1, None, None, '0 - 1', '0 - 1', None, 0, None),
 
            ('PLATELET COUNT', 'HAEMATOLOGY', 'Lakhs/Cumm', 1.5, 4.0, None,
             'Platelet count for clotting function',
             1.5, 4.0, 1.5, 4.0, 1.5, 4.0, '1.5 - 4.0', '1.5 - 4.0', '1.5 - 4.0', 0, None),
 
            ('PACKED CELL VALUE (PCV)', 'HAEMATOLOGY', '%', 40, 54, None,
             'Haematocrit — percentage of blood that is RBCs',
             38.3, 48.6, 35.5, 44.9, 35, 49, '38.3 - 48.6', '35.5 - 44.9', '35 - 49', 0, None),
 
            ('ESR', 'HAEMATOLOGY', 'mm/1hour', 0, 20, None,
             'Erythrocyte sedimentation rate — inflammation marker',
             0, 20, 0, 29, None, None, '0 - 20 (M)', '0 - 29 (F)', None, 0, None),
 
            ('MCV', 'HAEMATOLOGY', 'fL', 83, 101, None,
             'Mean corpuscular volume',
             83, 101, 83, 101, None, None, '83 - 101', '83 - 101', None, 0, None),
 
            ('MCH', 'HAEMATOLOGY', 'pg', 27, 32, None,
             'Mean corpuscular haemoglobin',
             27, 32, 27, 32, None, None, '27 - 32', '27 - 32', None, 0, None),
 
            ('MCHC', 'HAEMATOLOGY', 'gms%', 31, 34, None,
             'Mean corpuscular haemoglobin concentration',
             31, 34, 31, 34, None, None, '31 - 34', '31 - 34', None, 0, None),
 
            ('PERIPHERAL BLOOD SMEAR', 'HAEMATOLOGY', 'text', None, None, 'Normal',
             'Peripheral blood smear examination',
             None, None, None, None, None, None, 'Normal', 'Normal', 'Normal', 0, None),
 
            ('BLOOD GROUP & RH TYPING', 'HAEMATOLOGY', 'text', None, None, 'ABO and Rh blood group',
             'Blood group determination',
             None, None, None, None, None, None, None, None, None, 0, None),
 
            ('BLEEDING TIME', 'HAEMATOLOGY', 'Minutes', 1, 3, None,
             'Bleeding time',
             1, 3, 1, 3, None, None, '1 - 3', '1 - 3', None, 0, None),
 
            ('CLOTTING TIME', 'HAEMATOLOGY', 'Minutes', 4, 8, None,
             'Clotting time',
             4, 8, 4, 8, None, None, '4 - 8', '4 - 8', None, 0, None),
 
            # ── HORMONES ────────────────────────────────────────────────────────────
            ('TSH', 'HORMONES', 'µIU/mL', 0.35, 4.94, None,
             'Thyroid stimulating hormone',
             0.35, 4.94, 0.35, 4.94, 0.51, 5.0,
             '0.35 - 4.94\nHyperthyroidism: <0.1\nHypothyroidism: >20',
             '0.35 - 4.94\nHyperthyroidism: <0.1\nHypothyroidism: >20',
             '0.51 - 5.0\nHyperthyroidism: <0.1\nHypothyroidism: >20', 0,
             'Circulating TSH measurement has been used for screening for euthyroidism, and diagnosis for hyperthyroidism & hypothyroidism. Suppressed TSH (<0.01 µIU/mL) suggests hyperthyroidism. Elevated concentration (>7 µIU/mL) suggests hypothyroidism.'),
 
            ('T3 (TOTAL)', 'HORMONES', 'ng/dL', 58, 159, None,
             'Triiodothyronine total',
             58, 159, 58, 159, 105, 333, '58 - 159', '58 - 159', '105 - 333', 0, None),
 
            ('T4 (TOTAL)', 'HORMONES', 'µg/dL', 4.87, 11.72, None,
             'Thyroxine total',
             4.87, 11.72, 4.87, 11.72, 1.64, 5.6, '4.87 - 11.72', '4.87 - 11.72', '1.64 - 5.6', 0, None),
 
            ('FREE T3', 'HORMONES', 'ng/dL', 2.3, 3.5, None,
             'Free triiodothyronine',
             2.3, 3.5, 2.3, 3.5, None, None, '2.3 - 3.5', '2.3 - 3.5', None, 0, None),
 
            ('FREE T4', 'HORMONES', 'ng/dL', 0.8, 1.71, None,
             'Free thyroxine',
             0.8, 1.71, 0.8, 1.71, 0.8, 1.71, '0.8 - 1.71', '0.8 - 1.71', '0.8 - 1.71', 0, None),
 
            ('THYROID FUNCTION TEST', 'HORMONES', '', None, None, 'TSH + T3 + T4',
             'Complete thyroid panel',
             None, None, None, None, None, None, None, None, None, 0,
             'Circulating TSH measurement used for screening euthyroidism, diagnosing hyperthyroidism & hypothyroidism.'),
 
            ('FSH', 'HORMONES', 'mIU/mL', 1.24, 7.8, None,
             'Follicle stimulating hormone',
             1.24, 7.8, 4.5, 19.8, None, None, '1.24 - 7.8 (M)', '4.5 - 19.8 (F)', None, 0, None),
 
            ('LH', 'HORMONES', 'mIU/mL', 1.7, 2.2, None,
             'Luteinizing hormone',
             1.7, 2.2, 5.18, 26.53, None, None, '1.7 - 2.2 (M)', '5.18 - 26.53 (F)', None, 0, None),
 
            ('PROLACTIN', 'HORMONES', 'ng/mL', 2.5, 10.2, None,
             'Prolactin hormone',
             2.5, 10.2, 2.5, 10.2, None, None, '2.5 - 10.2', '2.5 - 10.2', None, 0, None),
 
            ('TESTOSTERONE', 'HORMONES', 'ng/dL', 160, 950, None,
             'Testosterone (male sex hormone)',
             160, 950, 40, 181, None, None, '160 - 950 (M)', '40 - 181 (F)', None, 0, None),
 
            # ── SEROLOGY ────────────────────────────────────────────────────────────
            ('WIDAL', 'SEROLOGY', 'Titre', None, None, 'Negative [<1/40]',
             'Typhoid fever antibody test',
             None, None, None, None, None, None, 'Negative [<1/40]', 'Negative [<1/40]', None, 0, None),
 
            ('DENGUE TEST', 'SEROLOGY', '', None, None, 'Negative',
             'Dengue NS1 Antigen + IgM + IgG',
             None, None, None, None, None, None, 'Negative', 'Negative', None, 0, None),
 
            ('CRP (C-REACTIVE PROTEIN)', 'SEROLOGY', 'mg/l', 0, 6, None,
             'C-Reactive Protein',
             0, 6, 0, 6, None, None, '0 - 6', '0 - 6', None, 500, None),
 
            ('RA FACTOR', 'SEROLOGY', 'IU/mL', 0, 14, None,
             'Rheumatoid arthritis factor',
             0, 14, 0, 14, None, None, '0 - 14', '0 - 14', None, 0, None),
 
            ('VDRL', 'SEROLOGY', '', None, None, 'Non Reactive',
             'Venereal disease research laboratory test',
             None, None, None, None, None, None, 'Non Reactive', 'Non Reactive', None, 0, None),
 
            ('HIV 1 & 2', 'SEROLOGY', '', None, None, 'Non Reactive',
             'HIV antibody screening',
             None, None, None, None, None, None, 'Non Reactive', 'Non Reactive', None, 0, None),
 
            ('HBsAg', 'SEROLOGY', '', None, None, 'Non Reactive: <1.00, Reactive: >=1.00',
             'Hepatitis B surface antigen',
             None, None, None, None, None, None, 'Non Reactive:<1.00\nReactive:>=1.00', 'Non Reactive:<1.00\nReactive:>=1.00', None, 0, None),
 
            ('HCV ANTIBODY', 'SEROLOGY', '', None, None, '<1.0 Non-reactive, =>1.0 Reactive',
             'Hepatitis C virus antibody',
             None, None, None, None, None, None, '<1.0 Non-reactive\n=>1.0 Reactive', '<1.0 Non-reactive\n=>1.0 Reactive', None, 0, None),
 
            ('A.S.O TITRE', 'SEROLOGY', 'IU/L', 0, 200, None,
             'Anti-streptolysin O titre',
             0, 200, 0, 200, 0, 200, '0 - 200', '0 - 200', '0 - 200', 500,
             'Demonstration of acute or recent streptococcal infection. Elevated values are consistent with an antecedent infection by group A streptococci. Elevated ASO titers found in about 85% of individuals with rheumatic fever.'),
 
            ('LEPTOSPIRAL TEST', 'SEROLOGY', '', None, None, 'Negative',
             'Leptospirosis antibody test',
             None, None, None, None, None, None, 'Negative', 'Negative', None, 0, None),
 
            ('TPHA', 'SEROLOGY', '', None, None, 'Non Reactive',
             'Treponema pallidum haemagglutination (syphilis)',
             None, None, None, None, None, None, 'Non Reactive', 'Non Reactive', None, 0,
             'Used for qualitative and Semi-Quantitative detection of Antibodies to Treponema pallidum. Aids in the diagnosis of patients where syphilis is suspected.'),
 
            ('IGE', 'SEROLOGY', 'IU/mL', 0, 100, None,
             'Immunoglobulin E — allergy screening',
             0, 100, 0, 100, None, None, '0 - 100', '0 - 100', None, 0,
             'Useful as an initial (screening) test for allergic disease. Serum levels of IgE are increased in many patients with allergic diseases, parasitic diseases, allergic bronchopulmonary aspergillosis.'),
 
            ('C.A.-125', 'SEROLOGY', 'U/mL', None, 35, '< 35',
             'Cancer antigen 125 — ovarian cancer marker',
             None, 35, None, 35, None, None, '< 35', '< 35', None, 0,
             'Cancer antigen 125 (CA 125) is a glycoprotein antigen. Elevated serum CA 125 levels are seen in many patients with cancer of the ovary and various other malignancies.'),
 
            ('PCR', 'SEROLOGY', '', None, None, 'Negative',
             'Polymerase Chain Reaction test',
             None, None, None, None, None, None, 'Negative', 'Negative', None, 0, None),
 
            # ── CLINICAL PATHOLOGY ──────────────────────────────────────────────────
            ('URINE EXAMINATION', 'CLINICAL PATHOLOGY', '', None, None, 'Physical, Chemical & Microscopical',
             'Complete urine analysis',
             None, None, None, None, None, None, None, None, None, 0, None),
 
            ('MOTION EXAMINATION', 'CLINICAL PATHOLOGY', '', None, None, 'Physical, Chemical & Microscopical',
             'Stool analysis',
             None, None, None, None, None, None, None, None, None, 0, None),
 
            ('SEMEN EXAMINATION', 'CLINICAL PATHOLOGY', '', None, None, 'Physical, Chemical & Microscopical',
             'Semen analysis',
             None, None, None, None, None, None, None, None, None, 0, None),
 
            # ── MICROBIOLOGY ─────────────────────────────────────────────────────────
            ('URINE CULTURE & SENSITIVITY', 'MICROBIOLOGY', '', None, None, None,
             'Urine culture and antibiotic sensitivity',
             None, None, None, None, None, None, None, None, None, 0, None),
 
            ('BLOOD CULTURE & SENSITIVITY', 'MICROBIOLOGY', '', None, None, None,
             'Blood culture and antibiotic sensitivity',
             None, None, None, None, None, None, None, None, None, 0, None),
 
            ('SPUTUM CULTURE & SENSITIVITY', 'MICROBIOLOGY', '', None, None, None,
             'Sputum culture and antibiotic sensitivity',
             None, None, None, None, None, None, None, None, None, 0, None),
 
            ('STOOL CULTURE & SENSITIVITY', 'MICROBIOLOGY', '', None, None, None,
             'Stool culture and antibiotic sensitivity',
             None, None, None, None, None, None, None, None, None, 0, None),
 
            ('PUS CULTURE & SENSITIVITY', 'MICROBIOLOGY', '', None, None, None,
             'Pus culture and antibiotic sensitivity',
             None, None, None, None, None, None, None, None, None, 0, None),
 
            # ── ENDOCRINOLOGY ────────────────────────────────────────────────────────
            ('INSULIN FASTING', 'ENDOCRINOLOGY', 'µIU/mL', 2.6, 3.7, None,
             'Fasting insulin level',
             2.6, 3.7, 2.6, 3.7, None, None, '2.6 - 3.7', '2.6 - 3.7', None, 0, None),
 
            ('CORTISOL', 'ENDOCRINOLOGY', 'µg/dL', 6, 38, None,
             'Cortisol — stress hormone',
             6, 38, 6, 38, None, None, '6 - 38', '6 - 38', None, 0, None),
 
            ('PARATHYROID HORMONE (PTH)', 'ENDOCRINOLOGY', 'pg/mL', 15, 65, None,
             'Parathyroid hormone',
             15, 65, 15, 65, None, None, '15 - 65', '15 - 65', None, 0, None),
 
            # ── OTHER TEST ───────────────────────────────────────────────────────────
            ('FERRITIN', 'OTHER TEST', 'ng/mL', 15, 200, None,
             'Iron storage protein',
             15, 200, 12, 150, None, None, '15 - 200 (M)', '12 - 150 (F)', None, 0, None),
 
            ('TIBC', 'OTHER TEST', 'mcg/dL', 250, 370, None,
             'Total iron binding capacity',
             250, 370, 250, 370, None, None, '250 - 370', '250 - 370', None, 0, None),
 
            ('PROTHROMBIN TIME', 'OTHER TEST', 'Seconds', 11, 17, 'Standard Therapy: 2.0 - 3.0',
             'Prothrombin time — coagulation test',
             11, 17, 11, 17, None, None,
             '11 - 17\nStandard Therapy: 2.0 - 3.0',
             '11 - 17\nStandard Therapy: 2.0 - 3.0', None, 0, None),
 
        # ── HAEMATOLOGY (new) ─────────────────────────────────────────────────
        ('PCV/HCT', 'HAEMATOLOGY', '%', None, None, 'M: 40-54%, F: 36-46%',
         'Packed Cell Volume / Haematocrit',
         40, 54, 36, 46, 33, 44,
         '40 - 54', '36 - 46', '33 - 44', 50, None),
 
        ('DIFFERENTIAL COUNT (DC)', 'HAEMATOLOGY', '%', None, None,
         'N: 40-70%, L: 20-40%, E: 0-6%, M: 0-2%, B: 0-1%',
         'Differential leucocyte count',
         None, None, None, None, None, None,
         'N:40-70 L:20-40 E:0-6 M:0-2 B:0-1',
         'N:40-70 L:20-40 E:0-6 M:0-2 B:0-1', None, 50, None),
 
        ('ABS EOSINOPHIL COUNT', 'HAEMATOLOGY', 'cells/cumm', 40, 400, None,
         'Absolute eosinophil count',
         40, 400, 40, 400, 40, 400, '40 - 400', '40 - 400', '40 - 400', 80, None),
 
        ('ABS NEUTROPHIL COUNT', 'HAEMATOLOGY', 'cells/cumm', 1800, 7500, None,
         'Absolute neutrophil count',
         1800, 7500, 1800, 7500, 1800, 7500,
         '1800 - 7500', '1800 - 7500', '1800 - 7500', 80, None),
 
        ('BONE MARROW STUDY', 'HAEMATOLOGY', '', None, None, None,
         'Bone marrow aspiration and biopsy examination',
         None, None, None, None, None, None, None, None, None, 0, None),
 
        ('MALARIA PARASITE (MP)', 'HAEMATOLOGY', '', None, None, 'Negative',
         'Peripheral blood smear for malaria parasite',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 150,
         'Presence of ring forms, trophozoites or schizonts indicates active malaria infection.'),
 
        ('COOMB\'S TEST (DIRECT)', 'HAEMATOLOGY', '', None, None, 'Negative',
         'Direct antiglobulin test — detects antibodies on RBC surface',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 150, None),
 
        ('COOMB\'S TEST (INDIRECT)', 'HAEMATOLOGY', '', None, None, 'Negative',
         'Indirect antiglobulin test — detects antibodies in serum',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 150, None),
 
        ('LE CELL', 'HAEMATOLOGY', '', None, None, 'Negative',
         'LE cell test for systemic lupus erythematosus',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 150, None),
 
        ('APTT', 'HAEMATOLOGY', 'seconds', 25, 35, None,
         'Activated Partial Thromboplastin Time — intrinsic coagulation pathway',
         25, 35, 25, 35, None, None, '25 - 35', '25 - 35', None, 150,
         'Prolonged APTT may indicate coagulation factor deficiency, heparin therapy, or DIC.'),
 
        ('BLOOD GAS ANALYSIS', 'HAEMATOLOGY', '', None, None, None,
         'Arterial blood gas analysis — pH, pO2, pCO2, HCO3',
         None, None, None, None, None, None, None, None, None, 0, None),
 
        # ── BIO-CHEMISTRY (new) ───────────────────────────────────────────────
        ('AMMONIA', 'BIO-CHEMISTRY', 'µmol/L', 11, 51, None,
         'Serum ammonia — liver function and hepatic encephalopathy marker',
         11, 51, 11, 51, 11, 51, '11 - 51', '11 - 51', '11 - 51', 250, None),
 
        ('CPK (CREATINE PHOSPHOKINASE)', 'BIO-CHEMISTRY', 'U/L', None, None,
         'M: 39-308 U/L, F: 26-192 U/L',
         'Creatine phosphokinase — muscle damage and cardiac marker',
         39, 308, 26, 192, None, None, '39 - 308', '26 - 192', None, 200, None),
 
        ('GCT (GLUCOSE CHALLENGE TEST)', 'BIO-CHEMISTRY', 'mg/dl', None, 140,
         'Normal: <140 mg/dl (1hr post 50g glucose)',
         'Glucose Challenge Test — gestational diabetes screening',
         None, 140, None, 140, None, None, None, '<140', None, 150, None),
 
        ('GGT (GAMMA GT)', 'BIO-CHEMISTRY', 'U/L', None, None,
         'M: 8-61 U/L, F: 5-36 U/L',
         'Gamma glutamyl transferase — liver disease and alcohol marker',
         8, 61, 5, 36, None, None, '8 - 61', '5 - 36', None, 150, None),
 
        ('GTT (GLUCOSE TOLERANCE TEST)', 'BIO-CHEMISTRY', 'mg/dl', None, None,
         'Fasting <100, 1hr <180, 2hr <140',
         'Oral Glucose Tolerance Test — diabetes diagnosis',
         None, None, None, None, None, None,
         'Fasting<100, 1hr<180, 2hr<140',
         'Fasting<100, 1hr<180, 2hr<140', None, 200, None),
 
        ('IGE (IMMUNOGLOBULIN E)', 'BIO-CHEMISTRY', 'IU/mL', None, 100,
         'Normal: <100 IU/mL',
         'Serum IgE — allergy and parasitic infection marker',
         None, 100, None, 100, None, None, '<100', '<100', None, 300, None),
 
        ('LACTATE (LACTIC ACID)', 'BIO-CHEMISTRY', 'mmol/L', 0.5, 2.2, None,
         'Serum lactate — tissue hypoxia and sepsis marker',
         0.5, 2.2, 0.5, 2.2, None, None, '0.5 - 2.2', '0.5 - 2.2', None, 250, None),
 
        ('LDH (LACTATE DEHYDROGENASE)', 'BIO-CHEMISTRY', 'U/L', 140, 280, None,
         'Lactate dehydrogenase — tissue damage marker',
         140, 280, 140, 280, None, None, '140 - 280', '140 - 280', None, 150, None),
 
        ('OSMOLALITY (SERUM)', 'BIO-CHEMISTRY', 'mOsm/kg', 275, 295, None,
         'Serum osmolality — fluid and electrolyte balance',
         275, 295, 275, 295, None, None, '275 - 295', '275 - 295', None, 200, None),
 
        ('OSMOLALITY (URINE)', 'BIO-CHEMISTRY', 'mOsm/kg', 300, 900, None,
         'Urine osmolality — kidney concentrating ability',
         300, 900, 300, 900, None, None, '300 - 900', '300 - 900', None, 200, None),
 
        ('VITAMIN B12', 'BIO-CHEMISTRY', 'pg/mL', 200, 900, None,
         'Serum Vitamin B12 — megaloblastic anaemia and neuropathy marker',
         200, 900, 200, 900, 200, 900, '200 - 900', '200 - 900', '200 - 900', 400,
         'Vitamin B12 <200 pg/mL indicates deficiency. Levels <100 pg/mL are associated with neurological symptoms.'),
 
        ('VLDL CHOLESTEROL', 'BIO-CHEMISTRY', 'mg/dl', 0, 40, 'Normal: <40 mg/dl (Calculated)',
         'Very Low Density Lipoprotein cholesterol',
         0, 40, 0, 40, None, None, '<40', '<40', None, 0, None),
 
        ('PROTEINS ELECTROPHORESIS', 'BIO-CHEMISTRY', '', None, None, None,
         'Serum protein electrophoresis — myeloma and dysproteinaemia screen',
         None, None, None, None, None, None, None, None, None, 500, None),
 
        ('LIPASE', 'BIO-CHEMISTRY', 'U/L', 13, 60, None,
         'Serum lipase — acute pancreatitis marker',
         13, 60, 13, 60, None, None, '13 - 60', '13 - 60', None, 200, None),
 
        ('% SATURATION (TRANSFERRIN)', 'BIO-CHEMISTRY', '%', 20, 50, None,
         'Transferrin saturation — iron status assessment',
         20, 50, 15, 50, None, None, '20 - 50', '15 - 50', None, 150, None),
 
        # ── LIVER FUNCTION TEST (new sub-params) ──────────────────────────────
        ('GLOBULIN', 'LIVER FUNCTION TEST', 'g/dL', 2.0, 3.5, None,
         'Serum globulin — infection and inflammation marker',
         2.0, 3.5, 2.0, 3.5, None, None, '2.0 - 3.5', '2.0 - 3.5', None, 0, None),
 
        ('A/G RATIO', 'LIVER FUNCTION TEST', '', 1.0, 2.2, None,
         'Albumin to globulin ratio — liver and nutritional status',
         1.0, 2.2, 1.0, 2.2, None, None, '1.0 - 2.2', '1.0 - 2.2', None, 0, None),
 
        ('BILIRUBIN DIRECT', 'LIVER FUNCTION TEST', 'mg/dl', 0.0, 0.3,
         'Normal: 0.0 - 0.3 mg/dl',
         'Direct (conjugated) bilirubin',
         0.0, 0.3, 0.0, 0.3, None, None, '0.0 - 0.3', '0.0 - 0.3', None, 0, None),
 
        ('BILIRUBIN INDIRECT', 'LIVER FUNCTION TEST', 'mg/dl', 0.2, 0.7,
         'Normal: 0.2 - 0.7 mg/dl',
         'Indirect (unconjugated) bilirubin',
         0.2, 0.7, 0.2, 0.7, None, None, '0.2 - 0.7', '0.2 - 0.7', None, 0, None),
 
        # ── RENAL FUNCTION TEST (new) ─────────────────────────────────────────
        ('24 HR URINE PROTEIN', 'RENAL FUNCTION TEST', 'mg/24hr', None, 150,
         'Normal: <150 mg/24hr',
         '24-hour urine protein — nephrotic syndrome and renal disease',
         None, 150, None, 150, None, None, '<150', '<150', None, 200, None),
 
        ('URINE ALBUMIN CREATININE RATIO', 'RENAL FUNCTION TEST', 'mg/g', None, 30,
         'Normal: <30 mg/g, Microalbuminuria: 30-300, Macroalbuminuria: >300',
         'Urine ACR — early diabetic nephropathy detection',
         None, 30, None, 30, None, None,
         'Normal<30, Micro:30-300, Macro>300',
         'Normal<30, Micro:30-300, Macro>300', None, 150,
         'ACR >30 mg/g indicates kidney damage. Early detection helps prevent progression to chronic kidney disease.'),
 
        # ── THYROID PROFILE (new) ─────────────────────────────────────────────
        ('FREE T3', 'THYROID PROFILE', 'pg/mL', 2.3, 4.2, None,
         'Free triiodothyronine — active thyroid hormone',
         2.3, 4.2, 2.3, 4.2, None, None, '2.3 - 4.2', '2.3 - 4.2', None, 200, None),
 
        ('FREE T4', 'THYROID PROFILE', 'ng/dL', 0.89, 1.76, None,
         'Free thyroxine — thyroid function marker',
         0.89, 1.76, 0.89, 1.76, None, None, '0.89 - 1.76', '0.89 - 1.76', None, 200, None),
 
        ('ANTI TPO', 'THYROID PROFILE', 'IU/mL', None, 35,
         'Normal: <35 IU/mL',
         'Anti-thyroid peroxidase antibody — Hashimoto\'s thyroiditis marker',
         None, 35, None, 35, None, None, '<35', '<35', None, 400,
         'Elevated Anti-TPO indicates autoimmune thyroid disease (Hashimoto\'s thyroiditis or Graves\' disease).'),
 
        ('ANTI TG', 'THYROID PROFILE', 'IU/mL', None, 115,
         'Normal: <115 IU/mL',
         'Anti-thyroglobulin antibody — autoimmune thyroid marker',
         None, 115, None, 115, None, None, '<115', '<115', None, 400, None),
 
        # ── SPECIAL TESTS / HORMONES ──────────────────────────────────────────
        ('AMH (ANTI-MULLERIAN HORMONE)', 'HORMONES', 'ng/mL', None, None,
         'Reproductive age F: 1.0-3.5 ng/mL',
         'Anti-Müllerian hormone — ovarian reserve marker',
         None, None, 1.0, 3.5, None, None, None, '1.0 - 3.5', None, 1200,
         'Low AMH indicates reduced ovarian reserve. Used in fertility assessment and IVF planning.'),
 
        ('ESTRADIOL (E2)', 'HORMONES', 'pg/mL', None, None,
         'F follicular: 20-350, ovulatory: 150-750, luteal: 30-450',
         'Serum estradiol — female reproductive hormone',
         None, None, None, None, None, None, None,
         'Follicular:20-350, Ovulatory:150-750, Luteal:30-450', None, 600, None),
 
        ('PROGESTERONE', 'HORMONES', 'ng/mL', None, None,
         'F luteal: 1.7-27.0, M: 0.2-1.4',
         'Serum progesterone — corpus luteum and pregnancy marker',
         0.2, 1.4, 1.7, 27.0, None, None, '0.2 - 1.4', '1.7 - 27.0 (luteal)', None, 500, None),
 
        ('AFP (ALPHA FETOPROTEIN)', 'HORMONES', 'ng/mL', None, 8.5,
         'Normal: <8.5 ng/mL',
         'Alpha fetoprotein — liver cancer and neural tube defect marker',
         None, 8.5, None, 8.5, None, None, '<8.5', '<8.5', None, 800,
         'Elevated AFP may indicate hepatocellular carcinoma, germ cell tumours, or neural tube defects in pregnancy.'),
 
        ('B-HCG (BETA HCG)', 'HORMONES', 'mIU/mL', None, 5,
         'Non-pregnant: <5 mIU/mL',
         'Beta human chorionic gonadotropin — pregnancy and tumour marker',
         None, 5, None, 5, None, None, None, '<5 (non-pregnant)', None, 500, None),
 
        ('CA 19-9', 'HORMONES', 'U/mL', None, 37,
         'Normal: <37 U/mL',
         'Cancer antigen 19-9 — pancreatic and GI cancer marker',
         None, 37, None, 37, None, None, '<37', '<37', None, 900,
         'CA 19-9 >37 U/mL may indicate pancreatic, colorectal, or biliary cancer. Also elevated in pancreatitis.'),
 
        ('CA 125', 'HORMONES', 'U/mL', None, 35,
         'Normal: <35 U/mL',
         'Cancer antigen 125 — ovarian cancer marker',
         None, 35, None, 35, None, None, '<35', '<35', None, 900,
         'CA 125 is primarily used to monitor ovarian cancer. Elevated levels also occur in endometriosis and pelvic infections.'),
 
        ('CEA (CARCINOEMBRYONIC ANTIGEN)', 'HORMONES', 'ng/mL', None, None,
         'Non-smoker: <3.0, Smoker: <5.0',
         'Carcinoembryonic antigen — colorectal and other cancer marker',
         None, 3.0, None, 3.0, None, None, '<3.0 (non-smoker)', '<3.0 (non-smoker)', None, 800,
         'CEA is used to monitor colorectal, breast, lung, and GI cancers. Not useful for initial diagnosis.'),
 
        ('PSA TOTAL', 'HORMONES', 'ng/mL', None, 4.0,
         'Normal: <4.0 ng/mL',
         'Prostate-specific antigen — prostate cancer screening',
         None, 4.0, None, None, None, None, '<4.0', None, None, 700,
         'PSA 4-10 ng/mL is borderline. PSA >10 ng/mL has higher risk of prostate cancer. Free/total PSA ratio helps differentiation.'),
 
        ('PRO-BNP (NT-PROBNP)', 'HORMONES', 'pg/mL', None, 125,
         'Normal: <125 pg/mL (<75yrs)',
         'N-terminal pro B-type natriuretic peptide — heart failure marker',
         None, 125, None, 125, None, None, '<125', '<125', None, 1200,
         'Elevated NT-proBNP indicates cardiac stress. Levels >125 pg/mL suggest heart failure.'),
 
        ('INSULIN FASTING', 'HORMONES', 'µIU/mL', 2.6, 24.9, None,
         'Fasting insulin — insulin resistance assessment',
         2.6, 24.9, 2.6, 24.9, None, None, '2.6 - 24.9', '2.6 - 24.9', None, 600, None),
 
        # ── MICROBIOLOGY / SEROLOGY (new) ─────────────────────────────────────
        ('ASO TEST (ANTISTREPTOLYSIN O)', 'SEROLOGY', 'IU/L', None, 200,
         'Normal: <200 IU/L',
         'Antistreptolysin O — streptococcal infection marker',
         None, 200, None, 200, None, None, '<200', '<200', None, 200,
         'Elevated ASO titre indicates recent Group A Streptococcus infection. Used to diagnose rheumatic fever.'),
 
        ('DENGUE IGM', 'SEROLOGY', '', None, None, 'Negative',
         'Dengue IgM antibody — acute dengue infection',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 300,
         'Dengue IgM appears 3-5 days after onset of fever and persists for 2-3 months.'),
 
        ('DENGUE NS1 ANTIGEN', 'SEROLOGY', '', None, None, 'Negative',
         'Dengue NS1 antigen — early acute dengue detection (days 1-5)',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 350,
         'NS1 antigen is detectable in the first 5 days of dengue fever, before IgM antibodies develop.'),
 
        ('HBSAG (HEPATITIS B SURFACE ANTIGEN)', 'SEROLOGY', '', None, None, 'Negative',
         'Hepatitis B surface antigen — active HBV infection marker',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 200,
         'Positive HBsAg indicates active Hepatitis B infection or carrier state.'),
 
        ('HCVAB (HEPATITIS C ANTIBODY)', 'SEROLOGY', '', None, None, 'Negative',
         'Hepatitis C antibody — HCV exposure marker',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 300,
         'Positive HCV antibody indicates exposure to Hepatitis C virus. Confirm with HCV RNA.'),
 
        ('HAV AG (HEPATITIS A ANTIGEN)', 'SEROLOGY', '', None, None, 'Negative',
         'Hepatitis A virus antigen / anti-HAV IgM — acute HAV infection',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 300, None),
 
        ('CHIKUNGUNYA IGM', 'SEROLOGY', '', None, None, 'Negative',
         'Chikungunya IgM antibody — acute chikungunya infection',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 400, None),
 
        ('H. PYLORI IGM', 'SEROLOGY', '', None, None, 'Negative',
         'Helicobacter pylori IgM — active H. pylori gastric infection',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 350,
         'H. pylori is associated with peptic ulcers and gastric cancer. Positive result warrants eradication therapy.'),
 
        ('SCRUB TYPHUS AB', 'SEROLOGY', '', None, None, 'Negative',
         'Scrub typhus IgM antibody — Orientia tsutsugamushi infection',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 400, None),
 
        ('HIV 1 & 2 (SCREENING)', 'SEROLOGY', '', None, None, 'Non-Reactive',
         'HIV 1 & 2 antibodies + p24 antigen combo test',
         None, None, None, None, None, None, 'Non-Reactive', 'Non-Reactive', 'Non-Reactive', 400,
         'Non-reactive result indicates no HIV antibodies or p24 antigen detected. Reactive result requires confirmatory Western Blot.'),
 
        ('LEPTOSPIRAL IGM', 'SEROLOGY', '', None, None, 'Negative',
         'Leptospira IgM antibody — leptospirosis screening',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 400, None),
 
        ('PAUL BUNNELL TEST', 'SEROLOGY', '', None, None, 'Negative',
         'Monospot / Paul-Bunnell test — infectious mononucleosis (EBV)',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 200, None),
 
        ('BRUCELLA AB', 'SEROLOGY', '', None, None, 'Negative',
         'Brucella antibody — brucellosis screening',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 400, None),
 
        ('ANA (ANTINUCLEAR ANTIBODY)', 'SEROLOGY', '', None, None, 'Negative',
         'Antinuclear antibody — autoimmune disease screening (SLE, RA)',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 600, None),
 
        ('ACLA IGG/IGM', 'SEROLOGY', '', None, None, 'Negative',
         'Anticardiolipin antibody IgG/IgM — antiphospholipid syndrome',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 700, None),
 
        ('APLA IGG/IGM', 'SEROLOGY', '', None, None, 'Negative',
         'Antiphospholipid antibody — thrombosis and recurrent pregnancy loss',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 700, None),
 
        ('ANTI DS-DNA', 'SEROLOGY', 'IU/mL', None, 30,
         'Normal: <30 IU/mL',
         'Anti-double stranded DNA — SLE specific marker',
         None, 30, None, 30, None, None, '<30', '<30', None, 700, None),
 
        ('C3 COMPLEMENT', 'SEROLOGY', 'mg/dL', 90, 180, None,
         'Complement component C3 — immune complex disease',
         90, 180, 90, 180, None, None, '90 - 180', '90 - 180', None, 500, None),
 
        ('KOH MOUNT FOR FUNGUS', 'MICROBIOLOGY', '', None, None, 'Negative',
         'KOH preparation for fungal elements',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 150, None),
 
        ('SMEAR FOR AFB', 'MICROBIOLOGY', '', None, None, 'Negative',
         'Acid Fast Bacilli smear — tuberculosis screening',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 200, None),
 
        # ── STOOL TESTS (new) ─────────────────────────────────────────────────
        ('STOOL OCCULT BLOOD', 'CLINICAL PATHOLOGY', '', None, None, 'Negative',
         'Faecal occult blood test — GI bleeding detection',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 100,
         'Positive occult blood may indicate colorectal cancer, polyps, peptic ulcer, or inflammatory bowel disease.'),
 
        ('STOOL ROUTINE EXAMINATION', 'CLINICAL PATHOLOGY', '', None, None, None,
         'Stool routine examination — consistency, colour, parasites, RBC, pus cells',
         None, None, None, None, None, None, None, None, None, 80, None),
 
        # ── URINE (new sub-params) ────────────────────────────────────────────
        ('URINE GLUCOSE', 'CLINICAL PATHOLOGY', '', None, None, 'Negative',
         'Urine glucose — diabetes and renal glycosuria screening',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 0, None),
 
        ('URINE BILE SALTS', 'CLINICAL PATHOLOGY', '', None, None, 'Negative',
         'Urine bile salts — liver disease marker',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 0, None),
 
        ('URINE BILE PIGMENT', 'CLINICAL PATHOLOGY', '', None, None, 'Negative',
         'Urine bile pigment — biliary obstruction marker',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 0, None),
 
        ('URINE UROBILINOGEN', 'CLINICAL PATHOLOGY', '', None, None, 'Normal',
         'Urine urobilinogen — haemolysis and liver function marker',
         None, None, None, None, None, None, 'Normal', 'Normal', 'Normal', 0, None),
 
        ('URINE KETONE BODIES', 'CLINICAL PATHOLOGY', '', None, None, 'Negative',
         'Urine ketone bodies — diabetic ketoacidosis and starvation marker',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 0, None),
 
        ('URINE NITRITE', 'CLINICAL PATHOLOGY', '', None, None, 'Negative',
         'Urine nitrite — urinary tract infection marker',
         None, None, None, None, None, None, 'Negative', 'Negative', 'Negative', 0, None),
 
        ('URINE PH', 'CLINICAL PATHOLOGY', '', 4.6, 8.0, '4.6 - 8.0',
         'Urine pH — acid-base balance and urinary tract assessment',
         4.6, 8.0, 4.6, 8.0, 4.6, 8.0, '4.6 - 8.0', '4.6 - 8.0', '4.6 - 8.0', 0, None),
 
        ('URINE SPECIFIC GRAVITY', 'CLINICAL PATHOLOGY', '', 1.003, 1.035,
         '1.003 - 1.035',
         'Urine specific gravity — kidney concentrating ability',
         1.003, 1.035, 1.003, 1.035, 1.003, 1.035,
         '1.003 - 1.035', '1.003 - 1.035', '1.003 - 1.035', 0, None),
 
        ('URINE CAST', 'CLINICAL PATHOLOGY', '/HPF', None, None, 'Nil',
         'Urine casts — renal tubular disease marker',
         None, None, None, None, None, None, 'Nil', 'Nil', 'Nil', 0, None),
 
        ('URINE CRYSTALS', 'CLINICAL PATHOLOGY', '/HPF', None, None, 'Nil',
         'Urine crystals — calculi and metabolic disorder marker',
         None, None, None, None, None, None, 'Nil', 'Nil', 'Nil', 0, None),
 
        ('URINE RBC', 'CLINICAL PATHOLOGY', '/HPF', None, None, 'Nil',
         'Urine red blood cells — haematuria detection',
         None, None, None, None, None, None, 'Nil', 'Nil', 'Nil', 0, None),
 
        ('URINE PUS CELLS', 'CLINICAL PATHOLOGY', '/HPF', None, None, '0 - 5',
         'Urine pus cells — urinary tract infection marker',
         None, 5, None, 5, None, 5, '0 - 5', '0 - 5', '0 - 5', 0, None),
 
        ('URINE EPITHELIAL CELLS', 'CLINICAL PATHOLOGY', '/HPF', None, None, 'Few',
         'Urine epithelial cells',
         None, None, None, None, None, None, 'Few', 'Few', 'Few', 0, None),
 
        ('PREGNANCY TEST (URINE/SERUM)', 'CLINICAL PATHOLOGY', '', None, None,
         'Negative (non-pregnant)',
         'hCG pregnancy test',
         None, None, None, None, None, None, None, 'Negative (non-pregnant)', None, 100, None),
 
        # ── ANDROLOGY ─────────────────────────────────────────────────────────
        ('SEMEN ROUTINE EXAMINATION', 'ANDROLOGY', '', None, None, None,
         'Semen analysis — volume, count, motility, morphology',
         None, None, None, None, None, None,
         'Volume≥1.5ml, Count≥16M/ml, Motility≥42%, Morphology≥4%',
         None, None, 400,
         'WHO 2021 reference values: volume ≥1.5ml, sperm concentration ≥16 million/ml, total motility ≥42%, normal morphology ≥4%.'),
 
        ('POST COITAL TEST', 'ANDROLOGY', '', None, None, None,
         'Post coital test — cervical mucus and sperm interaction',
         None, None, None, None, None, None, None, None, None, 300, None),
 
        # ── CSF / BODY FLUIDS ─────────────────────────────────────────────────
        ('CSF CELL COUNT (TC/DC)', 'CLINICAL PATHOLOGY', 'cells/cumm', None, None,
         'TC: 0-5 lymphocytes/cumm',
         'CSF total and differential cell count — meningitis diagnosis',
         None, None, None, None, None, None,
         'TC: 0-5 lymphocytes', 'TC: 0-5 lymphocytes', None, 300, None),
 
        ('CSF PROTEIN', 'CLINICAL PATHOLOGY', 'mg/dL', 15, 45, None,
         'CSF protein — BBB integrity and meningitis marker',
         15, 45, 15, 45, None, None, '15 - 45', '15 - 45', None, 200, None),
 
        ('CSF SUGAR', 'CLINICAL PATHOLOGY', 'mg/dL', 50, 80, None,
         'CSF glucose — bacterial meningitis and TB meningitis',
         50, 80, 50, 80, None, None, '50 - 80', '50 - 80', None, 200, None),
 
        ('CSF GRAM STAIN', 'MICROBIOLOGY', '', None, None, 'No organisms seen',
         'CSF Gram stain — bacterial meningitis diagnosis',
         None, None, None, None, None, None,
         'No organisms seen', 'No organisms seen', 'No organisms seen', 150, None),
 
        # ── GYNAECOLOGY PANELS ────────────────────────────────────────────────
        ('ANTENATAL PROFILE I', 'HEALTH PACKAGES', '', None, None, None,
         'Antenatal Profile I — Hb, Blood group, VDRL, HBsAg, HIV, Urine RE',
         None, None, None, None, None, None, None, None, None, 800, None),
 
        ('ANTENATAL PROFILE II', 'HEALTH PACKAGES', '', None, None, None,
         'Antenatal Profile II — CBC, Blood group, VDRL, HBsAg, HIV, RBS, TSH, Urine RE',
         None, None, None, None, None, None, None, None, None, 1200, None),
 
        ('INFERTILITY PROFILE', 'HEALTH PACKAGES', '', None, None, None,
         'Infertility Profile — FSH, LH, Prolactin, TSH, AMH, Testosterone',
         None, None, None, None, None, None, None, None, None, 2000, None),
 
        ('IVF PANEL (WIFE)', 'HEALTH PACKAGES', '', None, None, None,
         'IVF Panel Wife — FSH, LH, AMH, E2, Prolactin, TSH, HIV, HBsAg, VDRL',
         None, None, None, None, None, None, None, None, None, 3000, None),
 
        ('IVF PANEL (HUSBAND)', 'HEALTH PACKAGES', '', None, None, None,
         'IVF Panel Husband — Semen analysis, HIV, HBsAg, VDRL, Blood group',
         None, None, None, None, None, None, None, None, None, 2000, None),
 
        ('PIH PROFILE', 'HEALTH PACKAGES', '', None, None, None,
         'Pregnancy Induced Hypertension Profile — CBC, Renal function, LFT, Urine RE',
         None, None, None, None, None, None, None, None, None, 1000, None),
 
        # ── CARDIOLOGY PANELS ─────────────────────────────────────────────────
        ('ACS PANEL (ICCU PANEL)', 'HEALTH PACKAGES', '', None, None, None,
         'Acute Coronary Syndrome Panel — Troponin I, CK-MB, LDH, CPK, CBC, LFT',
         None, None, None, None, None, None, None, None, None, 1500, None),
 
        ('HEART FAILURE PANEL', 'HEALTH PACKAGES', '', None, None, None,
         'Heart Failure Panel — NT-proBNP, Troponin, Renal function, Electrolytes',
         None, None, None, None, None, None, None, None, None, 2000, None),
 
        ('CHEST PAIN PANEL', 'HEALTH PACKAGES', '', None, None, None,
         'Chest Pain Panel — Troponin I/T, CPK, LDH, ECG correlation',
         None, None, None, None, None, None, None, None, None, 1200, None),
 
        # ── ORTHOPAEDIC PANELS ────────────────────────────────────────────────
        ('ARTHRITIS PANEL 1', 'HEALTH PACKAGES', '', None, None, None,
         'Arthritis Panel 1 — RA Factor, CRP, ESR, Uric Acid, CBC',
         None, None, None, None, None, None, None, None, None, 700, None),
 
        ('ARTHRITIS PANEL 2', 'HEALTH PACKAGES', '', None, None, None,
         'Arthritis Panel 2 — ANA, Anti CCP, RA Factor, CRP, ESR, ASO',
         None, None, None, None, None, None, None, None, None, 1500, None),
 
        ('FRACTURE PANEL', 'HEALTH PACKAGES', '', None, None, None,
         'Fracture Panel — Calcium, Phosphorus, ALP, Vitamin D, PTH',
         None, None, None, None, None, None, None, None, None, 1200, None),
 
        ('OSTEOPOROSIS PANEL', 'HEALTH PACKAGES', '', None, None, None,
         'Osteoporosis Panel — Calcium, Phosphorus, ALP, Vitamin D, PTH, DEXA',
         None, None, None, None, None, None, None, None, None, 1500, None),
 
        # ── MISCELLANEOUS ─────────────────────────────────────────────────────
        ('TZANCK TEST', 'CLINICAL PATHOLOGY', '', None, None, None,
         'Tzanck smear — Herpes virus infection (HSV/VZV) diagnosis',
         None, None, None, None, None, None, None, None, None, 200, None),
 
        ('URINE CULTURE & SENSITIVITY', 'MICROBIOLOGY', '', None, None,
         'No growth',
         'Urine culture for bacterial identification and antibiotic sensitivity',
         None, None, None, None, None, None,
         'No growth', 'No growth', 'No growth', 300,
         'Culture positive result requires antibiotic sensitivity testing to guide treatment.'),
        ]
 
        c.executemany(
            '''INSERT INTO test_catalog
               (test_name, category, unit, normal_min, normal_max, normal_text, description,
                normal_min_m, normal_max_m, normal_min_f, normal_max_f,
                normal_min_c, normal_max_c, normal_text_m, normal_text_f, normal_text_c,
                amount, interpretation)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            tests
        )
    conn.commit()
    conn.close()
 
init_db()
init_patients_db()
init_doctors_db()
seed_test_catalog()
 
# ── Migrate: add sort_order to stage tables if missing ────────────────────────
def migrate_sort_order():
    conn = get_db()
    c = conn.cursor()
    for table in ('test_stage1', 'test_stage2', 'test_stage3'):
        cols = [row[1] for row in c.execute(f"PRAGMA table_info({table})").fetchall()]
        if 'sort_order' not in cols:
            c.execute(f"ALTER TABLE {table} ADD COLUMN sort_order INTEGER DEFAULT 0")
            c.execute(f"UPDATE {table} SET sort_order = id")
    conn.commit()
    conn.close()
 
migrate_sort_order()
 
def seed_stages():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM test_stage1")
    if c.fetchone()[0] > 0:
        conn.close()
        return
 
    # Stage 1 — Test Types
    stage1_data = [
        'BIO-CHEMISTRY', 'HAEMATOLOGY', 'HORMONES', 'CLINICAL PATHOLOGY',
        'SEROLOGY', 'MICROBIOLOGY', 'HEALTH PACKAGES', 'ENDOCRINOLOGY', 'OTHER TEST'
    ]
    for name in stage1_data:
        c.execute("INSERT OR IGNORE INTO test_stage1 (name) VALUES (?)", (name,))
 
    def s1(name):
        return c.execute("SELECT id FROM test_stage1 WHERE name=?", (name,)).fetchone()[0]
 
    # Stage 2 — Test Names mapped to Stage 1
    stage2_data = [
        # BIO-CHEMISTRY
        ('BIO-CHEMISTRY', 'LIPID PROFILE'),
        ('BIO-CHEMISTRY', 'DIABETIC PACKAGE'),
        ('BIO-CHEMISTRY', 'LIVER FUNCTION TEST'),
        ('BIO-CHEMISTRY', 'RENAL FUNCTION TEST'),
        ('BIO-CHEMISTRY', 'ELECTROLYTES, SERUM'),
        ('BIO-CHEMISTRY', 'BILIRUBIN'),
        ('BIO-CHEMISTRY', 'TOTAL BILIRUBIN'),
        ('BIO-CHEMISTRY', 'PROTHROMBIN TIME (PT)'),
        ('BIO-CHEMISTRY', 'BLOOD SUGAR FASTING'),
        ('BIO-CHEMISTRY', 'BLOOD SUGAR PP'),
        ('BIO-CHEMISTRY', 'HbA1c'),
        ('BIO-CHEMISTRY', 'URIC ACID'),
        ('BIO-CHEMISTRY', 'CREATININE'),
        ('BIO-CHEMISTRY', 'BLOOD UREA'),
        # HAEMATOLOGY
        ('HAEMATOLOGY', 'COMPLETE BLOOD COUNT (CBC)'),
        ('HAEMATOLOGY', 'ESR'),
        ('HAEMATOLOGY', 'PERIPHERAL BLOOD SMEAR'),
        ('HAEMATOLOGY', 'BLOOD GROUP & RH TYPING'),
        ('HAEMATOLOGY', 'BLEEDING TIME & CLOTTING TIME'),
        ('HAEMATOLOGY', 'RETICULOCYTE COUNT'),
        # HORMONES
        ('HORMONES', 'THYROID FUNCTION TEST'),
        ('HORMONES', 'THYROID PROFILE TEST'),
        ('HORMONES', 'FSH'),
        ('HORMONES', 'LH'),
        ('HORMONES', 'PROLACTIN'),
        ('HORMONES', 'TESTOSTERONE'),
        ('HORMONES', 'ESTRADIOL (E2)'),
        # CLINICAL PATHOLOGY
        ('CLINICAL PATHOLOGY', 'URINE EXAMINATION'),
        ('CLINICAL PATHOLOGY', 'MOTION EXAMINATION'),
        ('CLINICAL PATHOLOGY', 'SEMEN EXAMINATION'),
        # SEROLOGY
        ('SEROLOGY', 'WIDAL'),
        ('SEROLOGY', 'DENGUE TEST'),
        ('SEROLOGY', 'LEPTOSPIRAL TEST'),
        ('SEROLOGY', 'CRP (C-REACTIVE PROTEIN)'),
        ('SEROLOGY', 'RA FACTOR'),
        ('SEROLOGY', 'VDRL'),
        ('SEROLOGY', 'HIV 1 & 2'),
        ('SEROLOGY', 'HBsAg'),
        ('SEROLOGY', 'HCV ANTIBODY'),
        # MICROBIOLOGY
        ('MICROBIOLOGY', 'URINE CULTURE & SENSITIVITY'),
        ('MICROBIOLOGY', 'SPUTUM CULTURE & SENSITIVITY'),
        ('MICROBIOLOGY', 'BLOOD CULTURE & SENSITIVITY'),
        ('MICROBIOLOGY', 'STOOL CULTURE & SENSITIVITY'),
        ('MICROBIOLOGY', 'PUS CULTURE & SENSITIVITY'),
        # HEALTH PACKAGES
        ('HEALTH PACKAGES', 'HEALTHY MAN'),
        ('HEALTH PACKAGES', 'WELL WOMEN'),
        ('HEALTH PACKAGES', 'SENIOR CITIZEN PACKAGE'),
        ('HEALTH PACKAGES', 'CARDIAC RISK PACKAGE'),
        # ENDOCRINOLOGY
        ('ENDOCRINOLOGY', 'INSULIN FASTING'),
        ('ENDOCRINOLOGY', 'CORTISOL'),
        ('ENDOCRINOLOGY', 'DHEA-S'),
        ('ENDOCRINOLOGY', 'PARATHYROID HORMONE (PTH)'),
        # OTHER TEST
        ('OTHER TEST', 'VITAMIN D'),
        ('OTHER TEST', 'VITAMIN B12'),
        ('OTHER TEST', 'FERRITIN'),
        ('OTHER TEST', 'IRON STUDIES'),
        ('OTHER TEST', 'CALCIUM'),
        ('OTHER TEST', 'PHOSPHORUS'),
        ('OTHER TEST', 'MAGNESIUM'),
    ]
    for s1_name, s2_name in stage2_data:
        c.execute("INSERT INTO test_stage2 (stage1_id, name) VALUES (?,?)", (s1(s1_name), s2_name))
 
    def s2(name):
        row = c.execute("SELECT id FROM test_stage2 WHERE name=?", (name,)).fetchone()
        return row[0] if row else None
 
    # Stage 3 — Sub-parameters mapped to Stage 2
    stage3_data = [
        # Urine Examination
        ('URINE EXAMINATION', 'PHYSICAL EXAMINATION'),
        ('URINE EXAMINATION', 'CHEMICAL EXAMINATION'),
        ('URINE EXAMINATION', 'MICROSCOPICAL EXAMINATION'),
        ('URINE EXAMINATION', 'DEPOSIT'),
        # Motion Examination
        ('MOTION EXAMINATION', 'PHYSICAL EXAMINATION'),
        ('MOTION EXAMINATION', 'CHEMICAL EXAMINATION'),
        ('MOTION EXAMINATION', 'MICROSCOPICAL EXAMINATION'),
        ('MOTION EXAMINATION', 'MACROSCOPICAL EXAMINATION'),
        # Semen Examination
        ('SEMEN EXAMINATION', 'PHYSICAL EXAMINATION'),
        ('SEMEN EXAMINATION', 'MICROSCOPICAL EXAMINATION'),
        ('SEMEN EXAMINATION', 'CHEMICAL EXAMINATION'),
        ('SEMEN EXAMINATION', 'Differential Count'),
        # CBC
        ('COMPLETE BLOOD COUNT (CBC)', 'Differential Count'),
        # Peripheral Blood Smear
        ('PERIPHERAL BLOOD SMEAR', 'Differential Count'),
        ('PERIPHERAL BLOOD SMEAR', 'MICROSCOPICAL EXAMINATION'),
        # LFT sub-params
        ('LIVER FUNCTION TEST', 'SGOT (AST)'),
        ('LIVER FUNCTION TEST', 'SGPT (ALT)'),
        ('LIVER FUNCTION TEST', 'ALKALINE PHOSPHATASE'),
        ('LIVER FUNCTION TEST', 'TOTAL BILIRUBIN'),
        ('LIVER FUNCTION TEST', 'DIRECT BILIRUBIN'),
        ('LIVER FUNCTION TEST', 'INDIRECT BILIRUBIN'),
        ('LIVER FUNCTION TEST', 'TOTAL PROTEIN'),
        ('LIVER FUNCTION TEST', 'ALBUMIN'),
        ('LIVER FUNCTION TEST', 'GLOBULIN'),
        ('LIVER FUNCTION TEST', 'A/G RATIO'),
        # RFT sub-params
        ('RENAL FUNCTION TEST', 'BLOOD UREA'),
        ('RENAL FUNCTION TEST', 'SERUM CREATININE'),
        ('RENAL FUNCTION TEST', 'URIC ACID'),
        ('RENAL FUNCTION TEST', 'eGFR'),
        # Lipid Profile sub-params
        ('LIPID PROFILE', 'TOTAL CHOLESTEROL'),
        ('LIPID PROFILE', 'TRIGLYCERIDES'),
        ('LIPID PROFILE', 'HDL CHOLESTEROL'),
        ('LIPID PROFILE', 'LDL CHOLESTEROL'),
        ('LIPID PROFILE', 'VLDL CHOLESTEROL'),
        ('LIPID PROFILE', 'LDL/HDL RATIO'),
        # Thyroid sub-params
        ('THYROID FUNCTION TEST', 'TSH'),
        ('THYROID FUNCTION TEST', 'T3 (TOTAL)'),
        ('THYROID FUNCTION TEST', 'T4 (TOTAL)'),
        ('THYROID PROFILE TEST', 'TSH'),
        ('THYROID PROFILE TEST', 'FREE T3'),
        ('THYROID PROFILE TEST', 'FREE T4'),
        # Electrolytes
        ('ELECTROLYTES, SERUM', 'SODIUM'),
        ('ELECTROLYTES, SERUM', 'POTASSIUM'),
        ('ELECTROLYTES, SERUM', 'CHLORIDE'),
        ('ELECTROLYTES, SERUM', 'BICARBONATE'),
    ]
    for s2_name, s3_name in stage3_data:
        s2id = s2(s2_name)
        if s2id:
            c.execute("INSERT INTO test_stage3 (stage2_id, name) VALUES (?,?)", (s2id, s3_name))
 
    conn.commit()
    conn.close()
 
seed_stages()
 
# ── PAGES ─────────────────────────────────────────────────────────────────────
def sync_stage_hierarchy_from_catalog():
    """Keep stage1 and stage2 aligned with catalog categories and sub-categories."""
    conn = get_db()
    c = conn.cursor()

    categories = c.execute("""
        SELECT DISTINCT TRIM(category) AS category
        FROM test_catalog
        WHERE category IS NOT NULL AND TRIM(category) != ''
        ORDER BY category
    """).fetchall()

    for order, row in enumerate(categories):
        category_name = row["category"]
        if _stage_exclusion_exists(1, category=category_name):
            continue
        c.execute(
            "INSERT OR IGNORE INTO test_stage1 (name, sort_order) VALUES (?, ?)",
            (category_name, order)
        )

    stage1_rows = c.execute("SELECT id, name FROM test_stage1").fetchall()
    stage1_by_name = {row["name"]: row["id"] for row in stage1_rows}

    stage2_pairs = c.execute("""
        SELECT DISTINCT TRIM(category) AS category, TRIM(sub_category) AS sub_category
        FROM test_catalog
        WHERE category IS NOT NULL AND TRIM(category) != ''
          AND sub_category IS NOT NULL AND TRIM(sub_category) != ''
        ORDER BY category, sub_category
    """).fetchall()

    for order, row in enumerate(stage2_pairs):
        category_name = row["category"]
        subcategory_name = row["sub_category"]
        if _stage_exclusion_exists(1, category=category_name):
            continue
        if _stage_exclusion_exists(2, category=category_name, sub_category=subcategory_name):
            continue
        stage1_id = stage1_by_name.get(category_name)
        if not stage1_id:
            continue
        exists = c.execute(
            "SELECT id FROM test_stage2 WHERE stage1_id=? AND UPPER(TRIM(name))=UPPER(TRIM(?))",
            (stage1_id, subcategory_name)
        ).fetchone()
        if not exists:
            c.execute(
                "INSERT INTO test_stage2 (stage1_id, name, sort_order) VALUES (?, ?, ?)",
                (stage1_id, subcategory_name, order)
            )

    conn.commit()
    conn.close()

def _norm_stage_value(value):
    return (value or '').strip().upper()

def _stage_exclusion_exists(level, category='', sub_category='', test_name=''):
    conn = get_db()
    row = conn.execute(
        """
        SELECT 1
        FROM stage_exclusions
        WHERE level=? AND category=? AND sub_category=? AND test_name=?
        LIMIT 1
        """,
        (int(level), _norm_stage_value(category), _norm_stage_value(sub_category), _norm_stage_value(test_name))
    ).fetchone()
    conn.close()
    return bool(row)

def _add_stage_exclusion(level, category='', sub_category='', test_name=''):
    conn = get_db()
    conn.execute(
        """
        INSERT OR IGNORE INTO stage_exclusions (level, category, sub_category, test_name)
        VALUES (?, ?, ?, ?)
        """,
        (int(level), _norm_stage_value(category), _norm_stage_value(sub_category), _norm_stage_value(test_name))
    )
    conn.commit()
    conn.close()

def _clear_stage_exclusions(level, category='', sub_category='', test_name=''):
    conn = get_db()
    if int(level) == 1:
        conn.execute(
            "DELETE FROM stage_exclusions WHERE category=?",
            (_norm_stage_value(category),)
        )
    elif int(level) == 2:
        conn.execute(
            "DELETE FROM stage_exclusions WHERE level IN (2, 3) AND category=? AND sub_category=?",
            (_norm_stage_value(category), _norm_stage_value(sub_category))
        )
    else:
        conn.execute(
            "DELETE FROM stage_exclusions WHERE level=3 AND category=? AND sub_category=? AND test_name=?",
            (_norm_stage_value(category), _norm_stage_value(sub_category), _norm_stage_value(test_name))
        )
    conn.commit()
    conn.close()

@app.route('/')
def index():
    # Verify template exists
    template_path = os.path.join(TEMPLATE_DIR, "index.html")
    if not os.path.exists(template_path):
        return f"Error: Template not found at {template_path}", 500
    return render_template("index.html")
 
# ── PATIENTS ──────────────────────────────────────────────────────────────────
@app.route('/api/patients', methods=['GET'])
def get_patients():
    conn = get_patients_db()
    q = request.args.get('q', '')
    if q:
        rows = conn.execute("SELECT * FROM patients WHERE name LIKE ? OR phone LIKE ? ORDER BY created_at DESC",
                            (f'%{q}%', f'%{q}%')).fetchall()
    else:
        rows = conn.execute("SELECT * FROM patients ORDER BY created_at DESC").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])
 
@app.route('/api/patients', methods=['POST'])
def add_patient():
    data = request.json
    conn = get_patients_db()
    c = conn.cursor()
    c.execute("INSERT INTO patients (name,age,gender,phone,email,address,blood_group,greeting) VALUES (?,?,?,?,?,?,?,?)",
              (data['name'], data.get('age'), data.get('gender'), data.get('phone'),
               data.get('email'), data.get('address'), data.get('blood_group'), data.get('greeting')))
    pid = c.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'id': pid, 'message': 'Patient added'})
 
@app.route('/api/patients/<int:pid>', methods=['PUT'])
def update_patient(pid):
    data = request.json
    conn = get_patients_db()
    conn.execute("UPDATE patients SET name=?,age=?,gender=?,phone=?,email=?,address=?,blood_group=?,greeting=? WHERE id=?",
                 (data['name'], data.get('age'), data.get('gender'), data.get('phone'),
                  data.get('email'), data.get('address'), data.get('blood_group'), data.get('greeting'), pid))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Updated'})
 
@app.route('/api/patients/<int:pid>', methods=['DELETE'])
def delete_patient(pid):
    conn = get_patients_db()
    conn.execute("DELETE FROM patients WHERE id=?", (pid,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Deleted'})
 
# ── DOCTORS ───────────────────────────────────────────────────────────────────
@app.route('/api/patients/upsert-by-name', methods=['POST'])
def upsert_patient_by_name():
    """Save a manually-entered patient if not already in DB (matched by name)."""
    data = request.json
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Name required'}), 400
    conn = get_patients_db()
    existing = conn.execute("SELECT id FROM patients WHERE LOWER(name)=LOWER(?)", (name,)).fetchone()
    if existing:
        conn.close()
        return jsonify({'id': existing['id'], 'created': False})
    c = conn.cursor()
    c.execute("INSERT INTO patients (name,age,gender,phone,address) VALUES (?,?,?,?,?)",
              (name, data.get('age'), data.get('gender'), data.get('phone'), data.get('address')))
    pid = c.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'id': pid, 'created': True})
 
@app.route('/api/doctors', methods=['GET'])
def get_doctors():
    conn = get_doctors_db()
    rows = conn.execute("SELECT * FROM doctors ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])
 
@app.route('/api/doctors', methods=['POST'])
def add_doctor():
    data = request.json
    conn = get_doctors_db()
    c = conn.cursor()
    c.execute("INSERT INTO doctors (name,specialization,hospital,phone,email) VALUES (?,?,?,?,?)",
              (data['name'], data.get('specialization'), data.get('hospital'), data.get('phone'), data.get('email')))
    pid = c.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'id': pid, 'message': 'Doctor added'})
 
@app.route('/api/doctors/upsert-by-name', methods=['POST'])
def upsert_doctor_by_name():
    """Save a manually-entered doctor if not already in DB (matched by name)."""
    data = request.json
    name = (data.get('name') or '').strip()
    if not name or name == '—':
        return jsonify({'error': 'Name required'}), 400
    conn = get_doctors_db()
    existing = conn.execute("SELECT id FROM doctors WHERE LOWER(name)=LOWER(?)", (name,)).fetchone()
    if existing:
        conn.close()
        return jsonify({'id': existing['id'], 'created': False})
    c = conn.cursor()
    c.execute("INSERT INTO doctors (name) VALUES (?)", (name,))
    did = c.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'id': did, 'created': True})
 
@app.route('/api/doctors/<int:did>', methods=['DELETE'])
def delete_doctor(did):
    conn = get_doctors_db()
    conn.execute("DELETE FROM doctors WHERE id=?", (did,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Deleted'})
 
# ── PATIENTS: DOWNLOAD PDF / EXCEL / UPLOAD ───────────────────────────────────
 
@app.route('/api/patients/download/excel', methods=['GET'])
def download_patients_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    conn = get_patients_db()
    rows = conn.execute("SELECT id,name,age,gender,greeting,phone,email,address,blood_group,created_at FROM patients ORDER BY id").fetchall()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patients"
    headers = ['ID','Name','Age','Gender','Greeting','Phone','Email','Address','Blood Group','Registered At']
    hfont  = Font(bold=True, color='FFFFFF', size=11)
    hfill  = PatternFill('solid', fgColor='1A4A7A')
    altfill= PatternFill('solid', fgColor='EFF6FF')
    border = Border(left=Side(style='thin',color='BFDBFE'),right=Side(style='thin',color='BFDBFE'),
                    top=Side(style='thin',color='BFDBFE'),bottom=Side(style='thin',color='BFDBFE'))
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font=hfont; cell.fill=hfill; cell.border=border
        cell.alignment=Alignment(horizontal='center',vertical='center')
    for ri, row in enumerate(rows, 2):
        fill = altfill if ri%2==0 else None
        for ci, val in enumerate(tuple(row), 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border=border
            cell.alignment=Alignment(horizontal='left',vertical='center')
            if fill: cell.fill=fill
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(max((len(str(c.value or '')) for c in col), default=8)+4, 40)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buf, as_attachment=True, download_name=f'patients_{ts}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
 
@app.route('/api/patients/download/pdf', methods=['GET'])
def download_patients_pdf():
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    conn = get_patients_db()
    rows = conn.execute("SELECT id,name,age,gender,phone,blood_group,created_at FROM patients ORDER BY id").fetchall()
    conn.close()
    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=landscape(A4), leftMargin=12*mm, rightMargin=12*mm, topMargin=0*mm, bottomMargin=18*mm,
                            canvasmaker=add_lab_logo_canvas(doc))
    primary = colors.HexColor('#1A4A7A'); accent = colors.HexColor('#0d9488'); alt_row = colors.HexColor('#EFF6FF')
    styles = getSampleStyleSheet()
    title_s = ParagraphStyle('T', fontSize=16, textColor=primary, fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=2)
    sub_s   = ParagraphStyle('S', fontSize=9,  textColor=colors.grey, alignment=TA_CENTER, spaceAfter=6)
    story   = [Paragraph('Patients Report', title_s),
               Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y %I:%M %p')}  ·  Total: {len(rows)}", sub_s),
               HRFlowable(width='100%', thickness=2, color=accent, spaceAfter=8)]
    headers = ['ID','Name','Age','Gender','Phone','Blood Group','Registered At']
    col_w   = [15*mm, 60*mm, 18*mm, 22*mm, 38*mm, 28*mm, 45*mm]
    hdr_row = [Paragraph(f'<b>{h}</b>', ParagraphStyle('th', fontSize=8, textColor=colors.white, alignment=TA_CENTER)) for h in headers]
    data    = [hdr_row] + [[Paragraph(str(v or '—'), ParagraphStyle('td', fontSize=8, alignment=TA_LEFT)) for v in tuple(r)] for r in rows]
    t = Table(data, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),primary),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,alt_row]),
                            ('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#BFDBFE')),('VALIGN',(0,0),(-1,-1),'TOP'),('PADDING',(0,0),(-1,-1),4)]))
    story.append(t)
    def footer(canvas, doc):
        canvas.saveState(); canvas.setFont('Helvetica',7); canvas.setFillColor(colors.grey)
        canvas.line(12*mm,12*mm,landscape(A4)[0]-12*mm,12*mm)
        canvas.drawCentredString(landscape(A4)[0]/2,7*mm,f"Patients Report · {datetime.now().strftime('%d %b %Y')}")
        canvas.restoreState()
    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    out.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(out, as_attachment=True, download_name=f'patients_{ts}.pdf', mimetype='application/pdf')
 
@app.route('/api/patients/upload/excel', methods=['POST'])
def upload_patients_excel():
    """Import patients from an uploaded Excel file (.xlsx)."""
    import openpyxl
    file = request.files.get('file')
    if not file or not file.filename.lower().endswith('.xlsx'):
        return jsonify({'error': 'Please upload a .xlsx file'}), 400
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [str(c.value or '').strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {h: i for i, h in enumerate(headers)}
        def gv(row, key, alt=None):
            idx = col.get(key, col.get(alt)) if alt else col.get(key)
            return str(row[idx].value or '').strip() if idx is not None and row[idx].value is not None else None
        conn = get_patients_db(); c = conn.cursor(); inserted = 0; skipped = 0
        for row in ws.iter_rows(min_row=2):
            name = gv(row, 'name')
            if not name: skipped += 1; continue
            existing = conn.execute("SELECT id FROM patients WHERE LOWER(name)=LOWER(?)", (name,)).fetchone()
            if existing: skipped += 1; continue
            c.execute("INSERT INTO patients (name,age,gender,phone,email,address,blood_group) VALUES (?,?,?,?,?,?,?)",
                      (name, gv(row,'age'), gv(row,'gender'), gv(row,'phone'), gv(row,'email'),
                       gv(row,'address'), gv(row,'blood group','blood_group')))
            inserted += 1
        conn.commit(); conn.close()
        return jsonify({'message': f'{inserted} patients imported, {skipped} skipped (duplicate/empty)'})
    except Exception as e:
        return jsonify({'error': f'Import failed: {str(e)}'}), 500
 
@app.route('/api/patients/upload/db', methods=['POST'])
def upload_patients_db():
    """Restore patients from an uploaded patients.db file."""
    file = request.files.get('file')
    if not file or not file.filename.lower().endswith('.db'):
        return jsonify({'error': 'Please upload a valid .db file'}), 400
    import shutil, tempfile
    tmp = tempfile.mktemp(suffix='.db')
    file.save(tmp)
    try:
        test = sqlite3.connect(tmp)
        test.execute("SELECT COUNT(*) FROM patients")
        test.close()
        if os.path.exists(PATIENTS_DB_PATH):
            shutil.copy2(PATIENTS_DB_PATH, PATIENTS_DB_PATH + '.bak')
        shutil.move(tmp, PATIENTS_DB_PATH)
        return jsonify({'message': 'Patients database restored successfully'})
    except Exception as e:
        if os.path.exists(PATIENTS_DB_PATH + '.bak'):
            shutil.copy2(PATIENTS_DB_PATH + '.bak', PATIENTS_DB_PATH)
        return jsonify({'error': f'Restore failed: {str(e)}'}), 500
 
# ── DOCTORS: DOWNLOAD PDF / EXCEL / UPLOAD ────────────────────────────────────
 
@app.route('/api/doctors/download/excel', methods=['GET'])
def download_doctors_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    conn = get_doctors_db()
    rows = conn.execute("SELECT id,name,specialization,hospital,phone,email,created_at FROM doctors ORDER BY id").fetchall()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Doctors"
    headers = ['ID','Name','Specialization','Hospital','Phone','Email','Registered At']
    hfont  = Font(bold=True, color='FFFFFF', size=11)
    hfill  = PatternFill('solid', fgColor='0d9488')
    altfill= PatternFill('solid', fgColor='ECFDF5')
    border = Border(left=Side(style='thin',color='99f6e4'),right=Side(style='thin',color='99f6e4'),
                    top=Side(style='thin',color='99f6e4'),bottom=Side(style='thin',color='99f6e4'))
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font=hfont; cell.fill=hfill; cell.border=border
        cell.alignment=Alignment(horizontal='center',vertical='center')
    for ri, row in enumerate(rows, 2):
        fill = altfill if ri%2==0 else None
        for ci, val in enumerate(tuple(row), 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border=border; cell.alignment=Alignment(horizontal='left',vertical='center')
            if fill: cell.fill=fill
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(max((len(str(c.value or '')) for c in col), default=8)+4, 40)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buf, as_attachment=True, download_name=f'doctors_{ts}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
 
@app.route('/api/doctors/download/pdf', methods=['GET'])
def download_doctors_pdf():
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    conn = get_doctors_db()
    rows = conn.execute("SELECT id,name,specialization,hospital,phone,email FROM doctors ORDER BY id").fetchall()
    conn.close()
    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=landscape(A4), leftMargin=12*mm, rightMargin=12*mm, topMargin=0*mm, bottomMargin=18*mm,
                            canvasmaker=add_lab_logo_canvas(doc))
    primary = colors.HexColor('#0d9488'); accent = colors.HexColor('#1A4A7A'); alt_row = colors.HexColor('#ECFDF5')
    title_s = ParagraphStyle('T', fontSize=16, textColor=primary, fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=2)
    sub_s   = ParagraphStyle('S', fontSize=9,  textColor=colors.grey, alignment=TA_CENTER, spaceAfter=6)
    story   = [Paragraph('Doctors Report', title_s),
               Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y %I:%M %p')}  ·  Total: {len(rows)}", sub_s),
               HRFlowable(width='100%', thickness=2, color=accent, spaceAfter=8)]
    headers = ['ID','Name','Specialization','Hospital','Phone','Email']
    col_w   = [15*mm, 55*mm, 55*mm, 60*mm, 35*mm, 50*mm]
    hdr_row = [Paragraph(f'<b>{h}</b>', ParagraphStyle('th', fontSize=8, textColor=colors.white, alignment=TA_CENTER)) for h in headers]
    data    = [hdr_row] + [[Paragraph(str(v or '—'), ParagraphStyle('td', fontSize=8, alignment=TA_LEFT)) for v in tuple(r)] for r in rows]
    t = Table(data, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),primary),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,alt_row]),
                            ('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#99f6e4')),('VALIGN',(0,0),(-1,-1),'TOP'),('PADDING',(0,0),(-1,-1),4)]))
    story.append(t)
    def footer(canvas, doc):
        canvas.saveState(); canvas.setFont('Helvetica',7); canvas.setFillColor(colors.grey)
        canvas.line(12*mm,12*mm,landscape(A4)[0]-12*mm,12*mm)
        canvas.drawCentredString(landscape(A4)[0]/2,7*mm,f"Doctors Report · {datetime.now().strftime('%d %b %Y')}")
        canvas.restoreState()
    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    out.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(out, as_attachment=True, download_name=f'doctors_{ts}.pdf', mimetype='application/pdf')
 
@app.route('/api/doctors/upload/excel', methods=['POST'])
def upload_doctors_excel():
    """Import doctors from an uploaded Excel file (.xlsx)."""
    import openpyxl
    file = request.files.get('file')
    if not file or not file.filename.lower().endswith('.xlsx'):
        return jsonify({'error': 'Please upload a .xlsx file'}), 400
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [str(c.value or '').strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {h: i for i, h in enumerate(headers)}
        def gv(row, *keys):
            for key in keys:
                idx = col.get(key)
                if idx is not None and row[idx].value is not None:
                    return str(row[idx].value).strip()
            return None
        conn = get_doctors_db(); c = conn.cursor(); inserted = 0; skipped = 0
        for row in ws.iter_rows(min_row=2):
            name = gv(row, 'name')
            if not name: skipped += 1; continue
            existing = conn.execute("SELECT id FROM doctors WHERE LOWER(name)=LOWER(?)", (name,)).fetchone()
            if existing: skipped += 1; continue
            c.execute("INSERT INTO doctors (name,specialization,hospital,phone,email) VALUES (?,?,?,?,?)",
                      (name, gv(row,'specialization'), gv(row,'hospital'), gv(row,'phone'), gv(row,'email')))
            inserted += 1
        conn.commit(); conn.close()
        return jsonify({'message': f'{inserted} doctors imported, {skipped} skipped (duplicate/empty)'})
    except Exception as e:
        return jsonify({'error': f'Import failed: {str(e)}'}), 500
 
@app.route('/api/doctors/upload/db', methods=['POST'])
def upload_doctors_db():
    """Restore doctors from an uploaded doctors.db file."""
    file = request.files.get('file')
    if not file or not file.filename.lower().endswith('.db'):
        return jsonify({'error': 'Please upload a valid .db file'}), 400
    import shutil, tempfile
    tmp = tempfile.mktemp(suffix='.db')
    file.save(tmp)
    try:
        test = sqlite3.connect(tmp)
        test.execute("SELECT COUNT(*) FROM doctors")
        test.close()
        if os.path.exists(DOCTORS_DB_PATH):
            shutil.copy2(DOCTORS_DB_PATH, DOCTORS_DB_PATH + '.bak')
        shutil.move(tmp, DOCTORS_DB_PATH)
        return jsonify({'message': 'Doctors database restored successfully'})
    except Exception as e:
        if os.path.exists(DOCTORS_DB_PATH + '.bak'):
            shutil.copy2(DOCTORS_DB_PATH + '.bak', DOCTORS_DB_PATH)
        return jsonify({'error': f'Restore failed: {str(e)}'}), 500
 
 
# ── PATIENTS: IMPORT FROM PDF ─────────────────────────────────────────────────
 
@app.route('/api/patients/upload/pdf', methods=['POST'])
def upload_patients_pdf():
    """Import patients from a PDF exported by LabSoft (table-based extraction)."""
    file = request.files.get('file')
    if not file or not file.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'Please upload a valid .pdf file'}), 400
    try:
        import pdfplumber, re
        # Column names we recognise in the header row (lowercased)
        PATIENT_COLS = ['id', 'name', 'age', 'gender', 'phone', 'blood group', 'registered at']
 
        conn = get_patients_db()
        c = conn.cursor()
        inserted = 0
        skipped  = 0
        col_map  = {}  # header_text -> index
 
        with pdfplumber.open(file) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    if not table or len(table) < 2:
                        continue
                    # Detect header row: must contain 'name'
                    header = [str(cell or '').strip().lower() for cell in table[0]]
                    if 'name' not in header:
                        continue
                    # Build column index map
                    col_map = {h: i for i, h in enumerate(header)}
 
                    def gcol(row, *keys):
                        for k in keys:
                            idx = col_map.get(k)
                            if idx is not None and idx < len(row) and row[idx]:
                                v = str(row[idx]).strip()
                                if v and v != '—':
                                    return v
                        return None
 
                    for row in table[1:]:
                        name = gcol(row, 'name')
                        if not name:
                            skipped += 1
                            continue
                        # Skip if already exists
                        exists = conn.execute(
                            "SELECT id FROM patients WHERE LOWER(name)=LOWER(?)", (name,)
                        ).fetchone()
                        if exists:
                            skipped += 1
                            continue
                        age    = gcol(row, 'age')
                        gender = gcol(row, 'gender')
                        phone  = gcol(row, 'phone')
                        blood  = gcol(row, 'blood group')
                        c.execute(
                            "INSERT INTO patients (name,age,gender,phone,blood_group) VALUES (?,?,?,?,?)",
                            (name, age, gender, phone, blood)
                        )
                        inserted += 1
 
        conn.commit()
        conn.close()
        if inserted == 0 and skipped == 0:
            return jsonify({'error': 'No patient table found in PDF. Make sure you upload a LabSoft Patients PDF.'}), 400
        return jsonify({'message': f'{inserted} patients imported, {skipped} skipped (duplicate/empty/header)'})
    except Exception as e:
        return jsonify({'error': f'PDF import failed: {str(e)}'}), 500
 
 
# ── DOCTORS: IMPORT FROM PDF ──────────────────────────────────────────────────
 
@app.route('/api/doctors/upload/pdf', methods=['POST'])
def upload_doctors_pdf():
    """Import doctors from a PDF exported by LabSoft (table-based extraction)."""
    file = request.files.get('file')
    if not file or not file.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'Please upload a valid .pdf file'}), 400
    try:
        import pdfplumber
 
        conn = get_doctors_db()
        c = conn.cursor()
        inserted = 0
        skipped  = 0
 
        with pdfplumber.open(file) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    if not table or len(table) < 2:
                        continue
                    header = [str(cell or '').strip().lower() for cell in table[0]]
                    if 'name' not in header:
                        continue
                    col_map = {h: i for i, h in enumerate(header)}
 
                    def gcol(row, *keys):
                        for k in keys:
                            idx = col_map.get(k)
                            if idx is not None and idx < len(row) and row[idx]:
                                v = str(row[idx]).strip()
                                if v and v != '—':
                                    return v
                        return None
 
                    for row in table[1:]:
                        name = gcol(row, 'name')
                        if not name:
                            skipped += 1
                            continue
                        exists = conn.execute(
                            "SELECT id FROM doctors WHERE LOWER(name)=LOWER(?)", (name,)
                        ).fetchone()
                        if exists:
                            skipped += 1
                            continue
                        spec     = gcol(row, 'specialization')
                        hospital = gcol(row, 'hospital')
                        phone    = gcol(row, 'phone')
                        email    = gcol(row, 'email')
                        c.execute(
                            "INSERT INTO doctors (name,specialization,hospital,phone,email) VALUES (?,?,?,?,?)",
                            (name, spec, hospital, phone, email)
                        )
                        inserted += 1
 
        conn.commit()
        conn.close()
        if inserted == 0 and skipped == 0:
            return jsonify({'error': 'No doctor table found in PDF. Make sure you upload a LabSoft Doctors PDF.'}), 400
        return jsonify({'message': f'{inserted} doctors imported, {skipped} skipped (duplicate/empty/header)'})
    except Exception as e:
        return jsonify({'error': f'PDF import failed: {str(e)}'}), 500
 
 
@app.route('/api/reports', methods=['GET'])
def get_reports():
    conn = get_db_joined()
    q = request.args.get('q', '')
    pid = request.args.get('patient_id', '')
    date = request.args.get('date', '')
    
    query = '''SELECT r.*, p.name as patient_name, p.age, p.gender, p.greeting, p.phone as patient_phone,
                       d.name as doctor_name, d.specialization
               FROM reports r
               LEFT JOIN patients p ON r.patient_id = p.id
               LEFT JOIN doctors d ON r.doctor_id = d.id'''
    params = []
    filters = []
    if q:
        filters.append("(p.name LIKE ? OR r.report_title LIKE ?)")
        params += [f'%{q}%', f'%{q}%']
    if pid:
        filters.append("r.patient_id=?")
        params.append(pid)
    if date:
        filters.append("r.report_date = ?")
        params.append(date)
        
    if filters:
        query += ' WHERE ' + ' AND '.join(filters)
    query += ' ORDER BY r.created_at DESC'
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])
 
@app.route('/api/reports', methods=['POST'])
def create_report():
    data = request.json
    patient_id = data.get('patient_id')
    doctor_id = data.get('doctor_id')
    title = data.get('report_title', 'Test Report')
    date = data.get('report_date', datetime.now().strftime('%Y-%m-%d'))
    status = data.get('status', 'pending')
    
    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT INTO reports (patient_id, doctor_id, report_title, report_date, status) VALUES (?,?,?,?,?)",
              (patient_id, doctor_id, title, date, status))
    rid = c.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'id': rid, 'message': 'Report record created'})
 
@app.route('/api/reports/save-workflow', methods=['POST'])
def save_report_workflow():
    data = request.json
    p_name = (data.get('patient_name') or '').strip()
    d_name = (data.get('doctor_name') or '').strip()
    r_title = data.get('report_title', 'Lab Report')
    r_date = data.get('report_date', datetime.now().strftime('%Y-%m-%d'))
    html = data.get('html_content', '')
    p_id = data.get('patient_id')
    p_no = data.get('patient_no', '')
    d_id = data.get('doctor_id')
 
    # Resolve Patient ID if not provided (manual mode)
    if not p_id and p_name:
        conn_p = get_patients_db()
        existing_p = conn_p.execute("SELECT id FROM patients WHERE LOWER(name)=LOWER(?)", (p_name,)).fetchone()
        if existing_p:
            p_id = existing_p['id']
        else:
            c_p = conn_p.cursor()
            c_p.execute("INSERT INTO patients (name,age,gender,phone,address) VALUES (?,?,?,?,?)",
                        (p_name, data.get('patient_age'), data.get('patient_gender'), data.get('patient_phone'), data.get('patient_address')))
            p_id = c_p.lastrowid
            conn_p.commit()
        conn_p.close()
 
    # Resolve Doctor ID if not provided
    if not d_id and d_name and d_name != '—':
        conn_d = get_doctors_db()
        existing_d = conn_d.execute("SELECT id FROM doctors WHERE LOWER(name)=LOWER(?)", (d_name,)).fetchone()
        if existing_d:
            d_id = existing_d['id']
        else:
            c_d = conn_d.cursor()
            c_d.execute("INSERT INTO doctors (name) VALUES (?)", (d_name,))
            d_id = c_d.lastrowid
            conn_d.commit()
        conn_d.close()
 
    # Create Report Record
    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT INTO reports (patient_id, doctor_id, report_title, report_date, patient_no, html_content, status) VALUES (?,?,?,?,?,?,?)",
              (p_id, d_id, r_title, r_date, p_no, html, 'completed'))
    rid = c.lastrowid
    conn.commit()
    conn.close()
 
    # Generate real PDF using xhtml2pdf (pisa)
    out_path = os.path.join(UPLOAD_FOLDER, f"report_{rid}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf")
    sanitized_html = fix_html_for_pdf(html)
    with open(out_path, "wb") as f:
        pisa.CreatePDF(sanitized_html, dest=f)
    
    conn = get_db()
    conn.execute("UPDATE reports SET pdf_path=? WHERE id=?", (out_path, rid))
    conn.commit()
    conn.close()
 
    return jsonify({'id': rid, 'message': 'Report saved successfully'})
 
@app.route('/api/reports/<int:rid>', methods=['GET'])
def get_report(rid):
    conn = get_db_joined()
    row = conn.execute('''SELECT r.*, p.name as patient_name, p.age, p.gender, p.greeting,
                                 p.phone as patient_phone, p.email as patient_email,
                                 p.blood_group, d.name as doctor_name, d.specialization
                          FROM reports r
                          LEFT JOIN patients p ON r.patient_id = p.id
                          LEFT JOIN doctors d ON r.doctor_id = d.id
                          WHERE r.id=?''', (rid,)).fetchone()
    conn.close()
    if row:
        return jsonify(dict(row))
    return jsonify({'error': 'Not found'}), 404
 
@app.route('/api/reports/upload', methods=['POST'])
def upload_report():
    patient_id = request.form.get('patient_id')
    doctor_id = request.form.get('doctor_id')
    report_title = request.form.get('report_title', 'Lab Report')
    report_date = request.form.get('report_date', datetime.now().strftime('%Y-%m-%d'))
    file = request.files.get('pdf_file')
 
    extracted_text = ''
    original_pdf_path = ''
 
    if file and file.filename.endswith('.pdf'):
        fname = f"orig_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}"
        fpath = os.path.join(UPLOAD_FOLDER, fname)
        file.save(fpath)
        original_pdf_path = fpath
        # Extract text
        try:
            with pdfplumber.open(fpath) as pdf:
                for page in pdf.pages:
                    t = page.extract_text()
                    if t:
                        extracted_text += t + '\n'
        except Exception as e:
            extracted_text = f"Extraction error: {str(e)}"
 
    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT INTO reports (patient_id,doctor_id,report_title,report_date,extracted_text,original_pdf_path,status) VALUES (?,?,?,?,?,?,?)",
              (patient_id, doctor_id, report_title, report_date, extracted_text, original_pdf_path, 'uploaded'))
    rid = c.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'id': rid, 'extracted_text': extracted_text, 'message': 'Report uploaded'})
 
@app.route('/api/reports/<int:rid>', methods=['DELETE'])
def delete_report(rid):
    conn = get_db()
    conn.execute("DELETE FROM reports WHERE id=?", (rid,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Deleted'})
 
 
# ── PDF COMPATIBILITY UTILITIES ──────────────────────────────────────────────
PDF_FOOTER_HEIGHT_MM = 25

def fix_html_for_pdf(html):
    """
    Sanitizes HTML/CSS for xhtml2pdf (pisa).
    xhtml2pdf is very picky and crashes on modern CSS like calc().
    Also converts the position:fixed footer block into a proper xhtml2pdf
    @page frame so the footer/address bar appears on every PDF page.

    Key rules:
    - Footer (rpt-footer-bar + rpt-footer-addr) is placed in a fixed @frame
      that sits at the bottom of EVERY page — it never floats in the content flow.
    - The @page margin-bottom is set precisely to PDF_FOOTER_HEIGHT_MM (25mm)
      so content can never flow underneath the footer frame.
    - page-break-inside:avoid is kept on the footer block so it is never split.
    - break-inside is NOT stripped (xhtml2pdf honours page-break-inside instead,
      but we keep both so whichever engine picks up the CSS works correctly).
    - Tables and rows get page-break-inside:avoid to prevent splitting across footer.
    """
    if not html:
        return ""

    import re

    # 0. Extract footer HTML and convert to xhtml2pdf @page frame
    # The JS puts footer content inside .rpt-footer-block which contains
    # .rpt-footer-bar (coloured band) and .rpt-footer-addr (address line).
    # We capture the FULL inner content of rpt-footer-block (greedy inner match).
    footer_match = re.search(
        r'<div[^>]*class="rpt-footer-block"[^>]*>(.*?)</div\s*>(?=\s*(?:</div\s*>)?\s*</body)',
        html, re.DOTALL | re.IGNORECASE
    )
    if not footer_match:
        # Fallback: broader match — find rpt-footer-block and capture everything inside
        footer_match = re.search(
            r'<div[^>]*class=["\']rpt-footer-block["\'][^>]*>([\s\S]*?)</div>(?:\s*</div>)?\s*</body>',
            html, re.IGNORECASE
        )

    if footer_match:
        footer_inner = footer_match.group(1).strip()
        # Footer frame height: 30mm covers bar (~18mm) + address line (~8mm) + safety
        footer_frame_css = """
@page {
            size: A4 portrait;
            margin-top: 10mm;
            margin-bottom: 25mm;
            margin-left: 12mm;
            margin-right: 12mm;
            @frame footer_frame {
                -pdf-frame-content: pdf-footer-content;
                left: 12mm; right: 12mm;
                bottom: 0pt;
                height: 25mm;
            }
        }
        """
        html = re.sub(r'(<style[^>]*>)', r'\1' + footer_frame_css, html, count=1)

        # Build the repeating footer div (rendered into the frame on every page)
        footer_div = (
            f'<div id="pdf-footer-content" '
            f'style="width:100%;text-align:center;page-break-inside:avoid;">'
            f'{footer_inner}'
            f'</div>'
        )

        # Remove the spacer (it was only needed for screen layout)
        html = re.sub(r'<div[^>]*class="rpt-footer-spacer"[^>]*>\s*</div>', '', html)

        # Replace the original rpt-footer-block with our pdf-footer-content div.
        # Use a non-greedy match that stops at the FIRST closing </div> after the block opens.
        html = re.sub(
            r'<div[^>]*class=["\']rpt-footer-block["\'][^>]*>[\s\S]*?</div>',
            footer_div, html, count=1, flags=re.IGNORECASE
        )

    # 1. Remove calc() - pisa's parser (tinycss) crashes on it
    html = re.sub(r'calc\([^)]+\)', '100%', html)

    # 2. Convert rem/em to px (pisa doesn't support them well)
    def rem_to_px(match):
        val = float(match.group(1))
        return f"{int(val * 16)}px"
    html = re.sub(r'([0-9.]+)\s*rem', rem_to_px, html)
    html = re.sub(r'([0-9.]+)\s*em', rem_to_px, html)

    # 3. Remove display: flex / grid
    html = html.replace('display: flex;', 'display: block;')
    html = html.replace('display:flex;', 'display:block;')
    html = html.replace('display: grid;', 'display: block;')

    # 4. Remove CSS variables
    html = re.sub(r'var\([^)]+\)', 'inherit', html)

    # 5. Remove unsupported properties
    # NOTE: 'break-inside' and 'page-break-inside' are intentionally KEPT —
    # they prevent the footer/address bar from being split across pages.
    unsupported = ['box-sizing', 'pointer-events', 'z-index', 'backdrop-filter',
                   'object-fit', 'transition', 'transform', 'position']
    for prop in unsupported:
        html = re.sub(rf'\b{prop}\s*:[^;]+;', '', html)

    # 6. Fix width="100%" on img tags that crash pisa
    html = re.sub(r'<img([^>]+)width="100%"', r'<img\1width="450"', html)

    # 7. Remove modern CSS functions
    html = re.sub(r'(clamp|min|max|rgba|hsla)\([^)]+\)', 'inherit', html)

    # 8. Ensure every table row has page-break-inside:avoid so rows never
    #    straddle the footer boundary. Also add to tables themselves.
    html = re.sub(
        r'(<table\b[^>]*style=")',
        r'\1page-break-inside:avoid;',
        html, flags=re.IGNORECASE
    )
    html = re.sub(
        r'(<tr\b[^>]*style=")',
        r'\1page-break-inside:avoid; break-inside:avoid;',
        html, flags=re.IGNORECASE
    )
    html = re.sub(
        r'(<tr\b)(?![^>]*style=)([^>]*>)',
        r'\1 style="page-break-inside:avoid; break-inside:avoid;"\2',
        html, flags=re.IGNORECASE
    )

    return html
 
 
# ── GENERATE PDF REPORT ───────────────────────────────────────────────────────
@app.route('/api/reports/<int:rid>/generate-pdf', methods=['POST'])
def generate_pdf(rid):
    # Verify report exists
    conn = get_db()
    row = conn.execute('SELECT id FROM reports WHERE id=?', (rid,)).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'Not found'}), 404
 
    data = request.json or {}
    html_content = data.get('html_content', '')
    if not html_content:
        return jsonify({'error': 'No HTML content provided'}), 400
 
    out_path = os.path.join(UPLOAD_FOLDER, f"report_{rid}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf")
    pdf_generated = False
    
    try:
        # Sanitize HTML for pisa compatibility
        sanitized_html = fix_html_for_pdf(html_content)
        with open(out_path, "wb") as f:
            pisa.CreatePDF(sanitized_html, dest=f)
        pdf_generated = True
    except Exception as e:
        print(f"PDF render failed: {e}")
        out_path = None
 
    # Save HTML and path to DB (even if PDF failed, report is saved)
    conn = get_db()
    conn.execute("UPDATE reports SET html_content=?, pdf_path=?, status='completed' WHERE id=?", 
                 (html_content, out_path, rid))
    conn.commit()
    conn.close()
 
    return jsonify({
        'id': rid,
        'pdf_path': out_path, 
        'pdf_status': 'generated' if pdf_generated else 'failed (saved as HTML)',
        'message': 'Report saved successfully'
    })
 
 
 
 
 
 
# ── SETTINGS ──────────────────────────────────────────────────────────────────
@app.route('/api/settings', methods=['GET'])
def get_settings():
    conn = get_db()
    row = conn.execute("SELECT * FROM lab_settings LIMIT 1").fetchone()
    conn.close()
    return jsonify(dict(row) if row else {})
 
@app.route('/api/settings', methods=['POST'])
def save_settings():
    data = request.json
    conn = get_db()
    conn.execute('''UPDATE lab_settings SET lab_name=?, lab_address=?, lab_phone=?,
                    lab_email=?, form_design=?''',
                 (data.get('lab_name'), data.get('lab_address'), data.get('lab_phone'),
                  data.get('lab_email'), data.get('form_design')))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Settings saved'})
 
@app.route('/api/settings/logo', methods=['POST'])
def upload_logo():
    file = request.files.get('logo')
    if file:
        ext = os.path.splitext(file.filename)[1]
        fname = f"logo{ext}"
        fpath = os.path.join(LOGO_FOLDER, fname)
        file.save(fpath)
        conn = get_db()
        conn.execute("UPDATE lab_settings SET logo_path=?", (fpath,))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Logo uploaded', 'path': fpath})
    return jsonify({'error': 'No file'}), 400
 
# ── APPOINTMENTS ──────────────────────────────────────────────────────────────
@app.route('/api/appointments', methods=['GET'])
def get_appointments():
    conn = get_db_joined()
    rows = conn.execute('''
        SELECT a.*, p.name as patient_name, p.age, p.gender, p.greeting, p.phone as patient_phone,
               d.name as doctor_name
        FROM appointments a
        LEFT JOIN patients p ON a.patient_id = p.id
        LEFT JOIN doctors d ON a.doctor_id = d.id
        ORDER BY a.appointment_date DESC, a.appointment_time DESC
    ''').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])
 
@app.route('/api/appointments', methods=['POST'])
def add_appointment():
    data = request.json
    conn = get_db()
    c = conn.cursor()
    c.execute('''INSERT INTO appointments
                 (patient_id, doctor_id, appointment_date, appointment_time, test_names, notes, status)
                 VALUES (?,?,?,?,?,?,?)''',
              (data.get('patient_id'), data.get('doctor_id'),
               data.get('appointment_date'), data.get('appointment_time'),
               data.get('test_names'), data.get('notes'), data.get('status', 'scheduled')))
    aid = c.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'id': aid, 'message': 'Appointment booked'})
 
@app.route('/api/appointments/<int:aid>', methods=['PUT'])
def update_appointment(aid):
    data = request.json
    conn = get_db()
    conn.execute('''UPDATE appointments SET status=?, notes=? WHERE id=?''',
                 (data.get('status'), data.get('notes'), aid))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Updated'})
 
@app.route('/api/appointments/<int:aid>', methods=['DELETE'])
def delete_appointment(aid):
    conn = get_db()
    conn.execute("DELETE FROM appointments WHERE id=?", (aid,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Deleted'})
 
# ── TEST CATALOG ───────────────────────────────────────────────────────────────
@app.route('/api/tests', methods=['GET'])
def get_tests():
    conn = get_db()
    cat = request.args.get('category', '')
    q = request.args.get('q', '')
    if cat:
        rows = conn.execute("SELECT * FROM test_catalog WHERE category=? ORDER BY test_name", (cat,)).fetchall()
    elif q:
        rows = conn.execute("SELECT * FROM test_catalog WHERE test_name LIKE ? ORDER BY test_name", (f'%{q}%',)).fetchall()
    else:
        rows = conn.execute("SELECT * FROM test_catalog ORDER BY category, test_name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])
 
@app.route('/api/tests/by-name', methods=['GET'])
def get_test_by_name():
    name = request.args.get('name', '')
    conn = get_db()
    row = conn.execute("SELECT * FROM test_catalog WHERE test_name = ?", (name,)).fetchone()
    conn.close()
    if row:
        return jsonify(dict(row))
    return jsonify({}), 404
 
@app.route('/api/tests/categories', methods=['GET'])
def get_test_categories():
    conn = get_db()
    rows = conn.execute("SELECT DISTINCT category FROM test_catalog WHERE category IS NOT NULL AND TRIM(category) != '' ORDER BY category").fetchall()
    conn.close()
    return jsonify([r['category'] for r in rows])
 
@app.route('/api/tests/subcategories', methods=['GET'])
def get_test_subcategories():
    """Return distinct sub_category values from test_catalog, optionally filtered by category (stage1)."""
    category = request.args.get('category', '').strip().upper()
    conn = get_db()
    if category:
        rows = conn.execute(
            "SELECT DISTINCT category, sub_category FROM test_catalog WHERE sub_category IS NOT NULL AND TRIM(sub_category) != '' AND UPPER(TRIM(category)) = ? ORDER BY sub_category",
            (category,)
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT DISTINCT category, sub_category FROM test_catalog WHERE sub_category IS NOT NULL AND TRIM(sub_category) != '' ORDER BY category, sub_category"
        ).fetchall()
    conn.close()
    return jsonify([r['sub_category'] for r in rows])
 
@app.route('/api/tests/by-subcategory', methods=['GET'])
def get_tests_by_subcategory():
    """Return test_names from catalog filtered by category (stage1) and sub_category (stage2)."""
    category = request.args.get('category', '').strip().upper()
    sub_category = request.args.get('sub_category', '').strip().upper()
    conn = get_db()
    if category and sub_category:
        rows = conn.execute(
            "SELECT DISTINCT category, sub_category, test_name FROM test_catalog WHERE UPPER(TRIM(category)) = ? AND UPPER(TRIM(sub_category)) = ? ORDER BY test_name",
            (category, sub_category)
        ).fetchall()
    elif category:
        rows = conn.execute(
            "SELECT DISTINCT category, sub_category, test_name FROM test_catalog WHERE UPPER(TRIM(category)) = ? ORDER BY test_name",
            (category,)
        ).fetchall()
    else:
        rows = conn.execute("SELECT DISTINCT category, sub_category, test_name FROM test_catalog ORDER BY category, sub_category, test_name").fetchall()
    conn.close()
    return jsonify([r['test_name'] for r in rows])
 
@app.route('/api/tests/catalog-grouped', methods=['GET'])
def get_catalog_grouped():
    """Return test_catalog grouped by category for New Report entry and Bill entry.
    Each test includes unit, normal_min, normal_max, normal_text so the frontend
    can build reference-range strings without a second round-trip."""
    conn = get_db()
    rows = conn.execute(
        """SELECT test_name, category, sub_category, unit,
                  normal_min, normal_max, normal_text,
                  normal_min_m, normal_max_m, normal_text_m,
                  normal_min_f, normal_max_f, normal_text_f,
                  interpretation, amount
           FROM test_catalog
           ORDER BY category, test_name"""
    ).fetchall()
    conn.close()
 
    grouped = {}
    for r in rows:
        cat = r['category'] or 'MISCELLANEOUS'
        if cat not in grouped:
            grouped[cat] = []
 
        # Build a unified ref range string
        def ref_str(mn, mx, nt):
            if nt:
                return nt
            if mn is not None and mx is not None:
                return f"{mn} - {mx}"
            if mn is not None:
                return f"> {mn}"
            if mx is not None:
                return f"< {mx}"
            return ''
 
        ref = ref_str(r['normal_min'], r['normal_max'], r['normal_text'])
        ref_m = ref_str(r['normal_min_m'], r['normal_max_m'], r['normal_text_m'])
        ref_f = ref_str(r['normal_min_f'], r['normal_max_f'], r['normal_text_f'])
        sub_category = r['sub_category'] or ''

        grouped[cat].append({
            'name': r['test_name'],
            'unit': r['unit'] or '',
            'ref': ref,
            'ref_m': ref_m,
            'ref_f': ref_f,
            'interpretation': r['interpretation'] or '',
            'amount': r['amount'] or 0,
        })
 
    # Return as list of {category, tests[]} sorted by category
    return jsonify([
        {'category': cat, 'tests': tests}
        for cat, tests in sorted(grouped.items())
    ])
 
@app.route('/api/tests', methods=['POST'])
def add_test():
    data = request.json
    conn = get_db()
    c = conn.cursor()
    c.execute('''INSERT INTO test_catalog (test_name, category, sub_category, unit, normal_min, normal_max, normal_text, description)
                 VALUES (?,?,?,?,?,?,?,?)''',
              (data['test_name'], data.get('category'), data.get('sub_category') or None,
               data.get('unit'), data.get('normal_min'), data.get('normal_max'),
               data.get('normal_text'), data.get('description')))
    tid = c.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'id': tid, 'message': 'Test added'})
 
@app.route('/api/tests/<int:tid>', methods=['PUT'])
def update_test(tid):
    data = request.json
    conn = get_db()
    conn.execute('''UPDATE test_catalog SET test_name=?, category=?, sub_category=?, unit=?,
                    normal_min=?, normal_max=?, normal_text=?, description=?, interpretation=?
                    WHERE id=?''',
                 (data['test_name'], data.get('category'), data.get('sub_category'),
                  data.get('unit'), data.get('normal_min') or None, data.get('normal_max') or None,
                  data.get('normal_text'), data.get('description'),
                  data.get('interpretation'), tid))
    conn.commit(); conn.close()
    return jsonify({'message': 'Updated'})
 
@app.route('/api/tests/<int:tid>', methods=['DELETE'])
def delete_test(tid):
    conn = get_db()
    conn.execute('DELETE FROM test_catalog WHERE id=?', (tid,))
    conn.commit(); conn.close()
    return jsonify({'message': 'Deleted'})
 
@app.route('/api/stage1/<int:sid>', methods=['PUT'])
def update_stage1(sid):
    data = request.json
    conn = get_db()
    conn.execute("UPDATE test_stage1 SET name=? WHERE id=?", (data['name'].strip().upper(), sid))
    conn.commit(); conn.close()
    return jsonify({'message': 'Updated'})
 
@app.route('/api/stage2/<int:sid>', methods=['PUT'])
def update_stage2(sid):
    data = request.json
    conn = get_db()
    conn.execute("UPDATE test_stage2 SET name=? WHERE id=?", (data['name'].strip(), sid))
    conn.commit(); conn.close()
    return jsonify({'message': 'Updated'})
 
@app.route('/api/stage3/<int:sid>', methods=['PUT'])
def update_stage3(sid):
    data = request.json
    conn = get_db()
    conn.execute("UPDATE test_stage3 SET name=? WHERE id=?", (data['name'].strip(), sid))
    conn.commit(); conn.close()
    return jsonify({'message': 'Updated'})
 
# ── CATALOG / INTERPRETATIONS / STAGES EXPORT ─────────────────────────────────
 
@app.route('/api/export/catalog/<fmt>', methods=['GET'])
def export_catalog(fmt):
    conn = get_db()
    rows = conn.execute(
        "SELECT test_name,category,unit,normal_min,normal_max,normal_text,description,interpretation,amount FROM test_catalog ORDER BY category,test_name"
    ).fetchall()
    conn.close()
    headers = ['Test Name','Category','Unit','Normal Min','Normal Max','Normal Range','Description','Interpretation','Amount (₹)']
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
 
    if fmt == 'excel':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Test Catalog'
        hfont = Font(bold=True, color='FFFFFF'); hfill = PatternFill('solid', fgColor='1A4A7A')
        alt   = PatternFill('solid', fgColor='EFF6FF')
        thin  = Side(style='thin', color='BFDBFE')
        bdr   = Border(left=thin,right=thin,top=thin,bottom=thin)
        for ci,h in enumerate(headers,1):
            c=ws.cell(1,ci,h); c.font=hfont; c.fill=hfill; c.border=bdr
            c.alignment=Alignment(horizontal='center',vertical='center')
        for ri,row in enumerate(rows,2):
            f = alt if ri%2==0 else None
            for ci,val in enumerate(row,1):
                c=ws.cell(ri,ci,val); c.border=bdr
                c.alignment=Alignment(horizontal='left',vertical='top',wrap_text=True)
                if f: c.fill=f
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or '')) for c in col)+4,45)
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf,as_attachment=True,download_name=f'test_catalog_{timestamp}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
 
    # PDF
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, HRFlowable, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf,pagesize=landscape(A4),leftMargin=12*mm,rightMargin=12*mm,topMargin=14*mm,bottomMargin=18*mm)
    primary=colors.HexColor('#1A4A7A'); accent=colors.HexColor('#0d9488')
    alt_c=colors.HexColor('#EFF6FF'); styles=getSampleStyleSheet()
    title_s=ParagraphStyle('T',fontSize=16,textColor=primary,fontName='Helvetica-Bold',alignment=TA_CENTER,spaceAfter=4)
    sec_s=ParagraphStyle('S',fontSize=8,textColor=colors.grey,alignment=TA_CENTER,spaceAfter=8)
    th_s=ParagraphStyle('th',fontSize=8,textColor=colors.white,alignment=TA_CENTER,fontName='Helvetica-Bold')
    td_s=ParagraphStyle('td',fontSize=7.5,alignment=TA_LEFT,leading=10)
    story=[Paragraph('Test Catalog',title_s),
           Paragraph(f'Exported {datetime.now().strftime("%d %b %Y %I:%M %p")} · {len(rows)} tests',sec_s),
           HRFlowable(width='100%',thickness=2,color=accent,spaceAfter=8)]
    W=landscape(A4)[0]-24*mm
    hdr=[[Paragraph(h,th_s) for h in headers]]
    data_rows=[[Paragraph(str(v or '—'),td_s) for v in row] for row in rows]
    cw=[50*mm,30*mm,18*mm,18*mm,18*mm,38*mm,55*mm,55*mm,18*mm]
    t=Table(hdr+data_rows,colWidths=cw,repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),primary),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,alt_c]),
        ('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#BFDBFE')),
        ('VALIGN',(0,0),(-1,-1),'TOP'),('PADDING',(0,0),(-1,-1),4),
    ]))
    story.append(t)
    def footer(canvas,doc):
        canvas.saveState(); canvas.setFont('Helvetica',7); canvas.setFillColor(colors.grey)
        canvas.line(12*mm,12*mm,landscape(A4)[0]-12*mm,12*mm)
        canvas.drawCentredString(landscape(A4)[0]/2,7*mm,f'LabSoft · Test Catalog Export · {datetime.now().strftime("%d %b %Y")}')
        canvas.restoreState()
    doc.build(story,onFirstPage=footer,onLaterPages=footer)
    buf.seek(0)
    return send_file(buf,as_attachment=True,download_name=f'test_catalog_{timestamp}.pdf',mimetype='application/pdf')
 
 
@app.route('/api/export/interpretations/<fmt>', methods=['GET'])
def export_interpretations(fmt):
    conn = get_db()
    rows = conn.execute(
        "SELECT test_name,category,interpretation FROM test_catalog WHERE interpretation IS NOT NULL AND interpretation!='' ORDER BY category,test_name"
    ).fetchall()
    conn.close()
    headers = ['Test Name','Category','Interpretation']
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
 
    if fmt == 'excel':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb=openpyxl.Workbook(); ws=wb.active; ws.title='Interpretations'
        hfont=Font(bold=True,color='FFFFFF'); hfill=PatternFill('solid',fgColor='065F46')
        alt=PatternFill('solid',fgColor='ECFDF5')
        thin=Side(style='thin',color='A7F3D0'); bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
        for ci,h in enumerate(headers,1):
            c=ws.cell(1,ci,h); c.font=hfont; c.fill=hfill; c.border=bdr
            c.alignment=Alignment(horizontal='center',vertical='center')
        for ri,row in enumerate(rows,2):
            f=alt if ri%2==0 else None
            for ci,val in enumerate(row,1):
                c=ws.cell(ri,ci,val); c.border=bdr
                c.alignment=Alignment(horizontal='left',vertical='top',wrap_text=True)
                if f: c.fill=f
        ws.column_dimensions['A'].width=40; ws.column_dimensions['B'].width=22; ws.column_dimensions['C'].width=80
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf,as_attachment=True,download_name=f'interpretations_{timestamp}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
 
    # PDF
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    buf=io.BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=landscape(A4),leftMargin=12*mm,rightMargin=12*mm,topMargin=14*mm,bottomMargin=18*mm)
    primary=colors.HexColor('#065F46'); accent=colors.HexColor('#10b981'); alt_c=colors.HexColor('#ECFDF5')
    title_s=ParagraphStyle('T',fontSize=16,textColor=primary,fontName='Helvetica-Bold',alignment=TA_CENTER,spaceAfter=4)
    sub_s=ParagraphStyle('S',fontSize=8,textColor=colors.grey,alignment=TA_CENTER,spaceAfter=8)
    th_s=ParagraphStyle('th',fontSize=8,textColor=colors.white,alignment=TA_CENTER,fontName='Helvetica-Bold')
    td_s=ParagraphStyle('td',fontSize=8,alignment=TA_LEFT,leading=11)
    story=[Paragraph('Test Interpretations',title_s),
           Paragraph(f'Exported {datetime.now().strftime("%d %b %Y %I:%M %p")} · {len(rows)} entries',sub_s),
           HRFlowable(width='100%',thickness=2,color=accent,spaceAfter=8)]
    W=landscape(A4)[0]-24*mm
    hdr=[[Paragraph(h,th_s) for h in headers]]
    data_rows=[[Paragraph(str(v or '—'),td_s) for v in row] for row in rows]
    t=Table(hdr+data_rows,colWidths=[55*mm,30*mm,W-85*mm],repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),primary),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,alt_c]),
        ('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#A7F3D0')),
        ('VALIGN',(0,0),(-1,-1),'TOP'),('PADDING',(0,0),(-1,-1),5),
    ]))
    story.append(t)
    def footer(canvas,doc):
        canvas.saveState(); canvas.setFont('Helvetica',7); canvas.setFillColor(colors.grey)
        canvas.line(12*mm,12*mm,landscape(A4)[0]-12*mm,12*mm)
        canvas.drawCentredString(landscape(A4)[0]/2,7*mm,f'LabSoft · Test Interpretations Export · {datetime.now().strftime("%d %b %Y")}')
        canvas.restoreState()
    doc.build(story,onFirstPage=footer,onLaterPages=footer)
    buf.seek(0)
    return send_file(buf,as_attachment=True,download_name=f'interpretations_{timestamp}.pdf',mimetype='application/pdf')
 
 
@app.route('/api/export/stages/<fmt>', methods=['GET'])
def export_stages(fmt):
    conn = get_db()
    rows = conn.execute('''
        SELECT s1.name as stage1, s2.name as stage2, COALESCE(s3.name,'—') as stage3
        FROM test_stage2 s2
        JOIN test_stage1 s1 ON s2.stage1_id=s1.id
        LEFT JOIN test_stage3 s3 ON s3.stage2_id=s2.id
        ORDER BY s1.name, s2.name, s3.name
    ''').fetchall()
    conn.close()
    headers = ['Stage 1 — Test Type','Stage 2 — Test Name','Stage 3 — Sub Parameter']
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
 
    if fmt == 'excel':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb=openpyxl.Workbook(); ws=wb.active; ws.title='Test Stages'
        hfont=Font(bold=True,color='FFFFFF'); hfill=PatternFill('solid',fgColor='4C1D95')
        alt=PatternFill('solid',fgColor='F5F3FF')
        thin=Side(style='thin',color='C4B5FD'); bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
        for ci,h in enumerate(headers,1):
            c=ws.cell(1,ci,h); c.font=hfont; c.fill=hfill; c.border=bdr
            c.alignment=Alignment(horizontal='center',vertical='center')
        for ri,row in enumerate(rows,2):
            f=alt if ri%2==0 else None
            for ci,val in enumerate(row,1):
                c=ws.cell(ri,ci,val); c.border=bdr
                c.alignment=Alignment(horizontal='left',vertical='center')
                if f: c.fill=f
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width=45
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf,as_attachment=True,download_name=f'test_stages_{timestamp}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
 
    # PDF
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4
    buf=io.BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=A4,leftMargin=14*mm,rightMargin=14*mm,topMargin=14*mm,bottomMargin=18*mm)
    primary=colors.HexColor('#4C1D95'); accent=colors.HexColor('#8B5CF6'); alt_c=colors.HexColor('#F5F3FF')
    title_s=ParagraphStyle('T',fontSize=16,textColor=primary,fontName='Helvetica-Bold',alignment=TA_CENTER,spaceAfter=4)
    sub_s=ParagraphStyle('S',fontSize=8,textColor=colors.grey,alignment=TA_CENTER,spaceAfter=8)
    th_s=ParagraphStyle('th',fontSize=9,textColor=colors.white,alignment=TA_CENTER,fontName='Helvetica-Bold')
    td_s=ParagraphStyle('td',fontSize=8.5,alignment=TA_LEFT,leading=11)
    story=[Paragraph('Test Stages',title_s),
           Paragraph(f'Exported {datetime.now().strftime("%d %b %Y %I:%M %p")} · {len(rows)} entries',sub_s),
           HRFlowable(width='100%',thickness=2,color=accent,spaceAfter=8)]
    W=A4[0]-28*mm
    hdr=[[Paragraph(h,th_s) for h in headers]]
    data_rows=[[Paragraph(str(v or '—'),td_s) for v in row] for row in rows]
    t=Table(hdr+data_rows,colWidths=[W/3,W/3,W/3],repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),primary),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,alt_c]),
        ('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#C4B5FD')),
        ('VALIGN',(0,0),(-1,-1),'TOP'),('PADDING',(0,0),(-1,-1),5),
    ]))
    story.append(t)
    def footer(canvas,doc):
        canvas.saveState(); canvas.setFont('Helvetica',7); canvas.setFillColor(colors.grey)
        canvas.line(14*mm,12*mm,A4[0]-14*mm,12*mm)
        canvas.drawCentredString(A4[0]/2,7*mm,f'LabSoft · Test Stages Export · {datetime.now().strftime("%d %b %Y")}')
        canvas.restoreState()
    doc.build(story,onFirstPage=footer,onLaterPages=footer)
    buf.seek(0)
    return send_file(buf,as_attachment=True,download_name=f'test_stages_{timestamp}.pdf',mimetype='application/pdf')
 
 
    conn = get_db()
    conn.execute("DELETE FROM test_catalog WHERE id=?", (tid,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Deleted'})
 
# ── ENHANCED STATS ────────────────────────────────────────────────────────────
@app.route('/api/stats', methods=['GET'])
def get_stats():
    conn = get_db_joined()
    # Reports per day for last 7 days
    daily = conn.execute("""
        SELECT date(created_at) as day, COUNT(*) as count
        FROM reports
        WHERE created_at >= date('now', '-6 days')
        GROUP BY date(created_at)
        ORDER BY day
    """).fetchall()
    # Status breakdown
    status_counts = conn.execute("""
        SELECT status, COUNT(*) as count FROM reports GROUP BY status
    """).fetchall()
    # Appointments today
    appts_today = conn.execute(
        "SELECT COUNT(*) FROM appointments WHERE appointment_date=date('now')"
    ).fetchone()[0]
 
    stats = {
        'total_patients': conn.execute("SELECT COUNT(*) FROM patients").fetchone()[0],
        'total_reports': conn.execute("SELECT COUNT(*) FROM reports").fetchone()[0],
        'total_doctors': conn.execute("SELECT COUNT(*) FROM doctors").fetchone()[0],
        'reports_today': conn.execute("SELECT COUNT(*) FROM reports WHERE date(created_at)=date('now')").fetchone()[0],
        'appointments_today': appts_today,
        'daily_reports': [{'day': r['day'], 'count': r['count']} for r in daily],
        'status_breakdown': [{'status': r['status'], 'count': r['count']} for r in status_counts],
    }
    conn.close()
    return jsonify(stats)
 
# ── TEST STAGES API ───────────────────────────────────────────────────────────
 
@app.route('/api/stage1', methods=['GET'])
def get_stage1():
    sync_stage_hierarchy_from_catalog()
    conn = get_db()
    rows = conn.execute("SELECT * FROM test_stage1 ORDER BY sort_order, id").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])
 
@app.route('/api/stage1', methods=['POST'])
def add_stage1():
    data = request.json
    category_name = data['name'].strip().upper()
    _clear_stage_exclusions(1, category=category_name)
    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT OR IGNORE INTO test_stage1 (name) VALUES (?)", (category_name,))
    conn.commit()
    new_id = c.lastrowid
    if not new_id:
        row = conn.execute("SELECT id FROM test_stage1 WHERE name=?", (category_name,)).fetchone()
        new_id = row['id'] if row else None
    conn.close()
    return jsonify({'message': 'Stage 1 added', 'id': new_id})
 
@app.route('/api/stage1/<int:sid>', methods=['DELETE'])
def delete_stage1(sid):
    conn = get_db()
    row = conn.execute("SELECT name FROM test_stage1 WHERE id=?", (sid,)).fetchone()
    if row:
        category_name = row['name']
        catalog_match = conn.execute(
            "SELECT 1 FROM test_catalog WHERE UPPER(TRIM(category)) = UPPER(TRIM(?)) LIMIT 1",
            (category_name,)
        ).fetchone()
        if catalog_match:
            _add_stage_exclusion(1, category=category_name)
        stage2_rows = conn.execute("SELECT id, name FROM test_stage2 WHERE stage1_id=?", (sid,)).fetchall()
        for s2 in stage2_rows:
            if catalog_match:
                _add_stage_exclusion(2, category=category_name, sub_category=s2['name'])
            stage3_rows = conn.execute("SELECT id, name FROM test_stage3 WHERE stage2_id=?", (s2['id'],)).fetchall()
            for s3 in stage3_rows:
                if catalog_match:
                    _add_stage_exclusion(3, category=category_name, sub_category=s2['name'], test_name=s3['name'])
            conn.execute("DELETE FROM test_stage3 WHERE stage2_id=?", (s2['id'],))
        conn.execute("DELETE FROM test_stage2 WHERE stage1_id=?", (sid,))
    conn.execute("DELETE FROM test_stage1 WHERE id=?", (sid,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Deleted'})
 
@app.route('/api/stage2', methods=['GET'])
def get_stage2():
    sync_stage_hierarchy_from_catalog()
    s1 = request.args.get('stage1_id', '')
    conn = get_db()
    if s1:
        rows = conn.execute(
            "SELECT s2.*, s1.name as stage1_name FROM test_stage2 s2 JOIN test_stage1 s1 ON s2.stage1_id=s1.id WHERE s2.stage1_id=? ORDER BY s2.sort_order, s2.id",
            (s1,)).fetchall()
    else:
        rows = conn.execute(
            "SELECT s2.*, s1.name as stage1_name FROM test_stage2 s2 JOIN test_stage1 s1 ON s2.stage1_id=s1.id ORDER BY s1.sort_order, s1.id, s2.sort_order, s2.id"
        ).fetchall()
    conn.close()
    seen = set()
    payload = []
    for r in rows:
        key = f"{r['stage1_id']}::{(r['name'] or '').strip().lower()}"
        if key in seen:
            continue
        seen.add(key)
        payload.append(dict(r))
    return jsonify(payload)
 
@app.route('/api/stage2', methods=['POST'])
def add_stage2():
    data = request.json
    subcategory_name = data['name'].strip()
    conn = get_db()
    parent = conn.execute("SELECT name FROM test_stage1 WHERE id=?", (data['stage1_id'],)).fetchone()
    if parent:
        _clear_stage_exclusions(2, category=parent['name'], sub_category=subcategory_name)
    c = conn.cursor()
    existing = conn.execute(
        "SELECT id FROM test_stage2 WHERE stage1_id=? AND UPPER(TRIM(name))=UPPER(TRIM(?))",
        (data['stage1_id'], subcategory_name)
    ).fetchone()
    if existing:
        new_id = existing['id']
    else:
        c.execute("INSERT INTO test_stage2 (stage1_id, name) VALUES (?,?)",
                  (data['stage1_id'], subcategory_name))
        new_id = c.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'message': 'Stage 2 added', 'id': new_id})
 
@app.route('/api/stage2/<int:sid>', methods=['DELETE'])
def delete_stage2(sid):
    conn = get_db()
    row = conn.execute("""
        SELECT s2.name AS stage2_name, s1.name AS stage1_name
        FROM test_stage2 s2
        JOIN test_stage1 s1 ON s2.stage1_id = s1.id
        WHERE s2.id=?
    """, (sid,)).fetchone()
    if row:
        matching_rows = conn.execute(
            """
            SELECT id, name
            FROM test_stage2
            WHERE stage1_id=(
                SELECT stage1_id FROM test_stage2 WHERE id=?
            )
              AND UPPER(TRIM(name)) = UPPER(TRIM(?))
            """,
            (sid, row['stage2_name'])
        ).fetchall()
        catalog_match = conn.execute(
            """
            SELECT 1
            FROM test_catalog
            WHERE UPPER(TRIM(category)) = UPPER(TRIM(?))
              AND UPPER(TRIM(sub_category)) = UPPER(TRIM(?))
            LIMIT 1
            """,
            (row['stage1_name'], row['stage2_name'])
        ).fetchone()
        if catalog_match:
            _add_stage_exclusion(2, category=row['stage1_name'], sub_category=row['stage2_name'])
        for s2row in matching_rows:
            stage3_rows = conn.execute("SELECT id, name FROM test_stage3 WHERE stage2_id=?", (s2row['id'],)).fetchall()
            for s3 in stage3_rows:
                if catalog_match:
                    _add_stage_exclusion(3, category=row['stage1_name'], sub_category=row['stage2_name'], test_name=s3['name'])
            conn.execute("DELETE FROM test_stage3 WHERE stage2_id=?", (s2row['id'],))
        conn.execute(
            """
            DELETE FROM test_stage2
            WHERE stage1_id=(
                SELECT stage1_id FROM test_stage2 WHERE id=?
            )
              AND UPPER(TRIM(name)) = UPPER(TRIM(?))
            """,
            (sid, row['stage2_name'])
        )
    conn.commit()
    conn.close()
    return jsonify({'message': 'Deleted'})
 
@app.route('/api/stage3', methods=['GET'])
def get_stage3():
    sync_stage_hierarchy_from_catalog()
    s2 = request.args.get('stage2_id', '')
    conn = get_db()
    if s2:
        rows = conn.execute(
            "SELECT s3.*, s2.name as stage2_name FROM test_stage3 s3 JOIN test_stage2 s2 ON s3.stage2_id=s2.id WHERE s3.stage2_id=? ORDER BY s3.sort_order, s3.id",
            (s2,)).fetchall()
    else:
        rows = conn.execute(
            "SELECT s3.*, s2.name as stage2_name FROM test_stage3 s3 JOIN test_stage2 s2 ON s3.stage2_id=s2.id ORDER BY s2.sort_order, s2.id, s3.sort_order, s3.id"
        ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])
 
@app.route('/api/stage3', methods=['POST'])
def add_stage3():
    data = request.json
    test_name = data['name'].strip()
    conn = get_db()
    parent = conn.execute("""
        SELECT s1.name AS stage1_name, s2.name AS stage2_name
        FROM test_stage2 s2
        JOIN test_stage1 s1 ON s2.stage1_id = s1.id
        WHERE s2.id=?
    """, (data['stage2_id'],)).fetchone()
    if parent:
        _clear_stage_exclusions(3, category=parent['stage1_name'], sub_category=parent['stage2_name'], test_name=test_name)
    c = conn.cursor()
    c.execute("INSERT INTO test_stage3 (stage2_id, name) VALUES (?,?)",
              (data['stage2_id'], test_name))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Stage 3 added'})
 
@app.route('/api/stage3/<int:sid>', methods=['DELETE'])
def delete_stage3(sid):
    conn = get_db()
    row = conn.execute("""
        SELECT s3.name AS stage3_name, s2.name AS stage2_name, s1.name AS stage1_name
        FROM test_stage3 s3
        JOIN test_stage2 s2 ON s3.stage2_id = s2.id
        JOIN test_stage1 s1 ON s2.stage1_id = s1.id
        WHERE s3.id=?
    """, (sid,)).fetchone()
    if row:
        catalog_match = conn.execute(
            """
            SELECT 1
            FROM test_catalog
            WHERE UPPER(TRIM(category)) = UPPER(TRIM(?))
              AND UPPER(TRIM(sub_category)) = UPPER(TRIM(?))
              AND UPPER(TRIM(test_name)) = UPPER(TRIM(?))
            LIMIT 1
            """,
            (row['stage1_name'], row['stage2_name'], row['stage3_name'])
        ).fetchone()
        if catalog_match:
            _add_stage_exclusion(3, category=row['stage1_name'], sub_category=row['stage2_name'], test_name=row['stage3_name'])
    conn.execute("DELETE FROM test_stage3 WHERE id=?", (sid,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Deleted'})
 
# ── STAGE REORDER ─────────────────────────────────────────────────────────────
@app.route('/api/stage1/reorder', methods=['PUT'])
def reorder_stage1():
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    conn = get_db()
    for order, sid in enumerate(ids):
        conn.execute("UPDATE test_stage1 SET sort_order=? WHERE id=?", (order, sid))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Reordered'})
 
@app.route('/api/stage2/reorder', methods=['PUT'])
def reorder_stage2():
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    conn = get_db()
    for order, sid in enumerate(ids):
        conn.execute("UPDATE test_stage2 SET sort_order=? WHERE id=?", (order, sid))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Reordered'})
 
@app.route('/api/stage3/reorder', methods=['PUT'])
def reorder_stage3():
    data = request.get_json(silent=True) or {}
    ids = data.get('ids', [])
    conn = get_db()
    for order, sid in enumerate(ids):
        conn.execute("UPDATE test_stage3 SET sort_order=? WHERE id=?", (order, sid))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Reordered'})
 
# ── BILLING ───────────────────────────────────────────────────────────────────
 
@app.route('/api/bills', methods=['GET'])
def get_bills():
    conn = get_db_joined()
    rows = conn.execute('''
        SELECT b.*, p.name as patient_name, p.age, p.gender, p.greeting, p.phone as patient_phone,
               d.name as doctor_name
        FROM bills b
        LEFT JOIN patients p ON b.patient_id = p.id
        LEFT JOIN doctors d ON b.doctor_id = d.id
        ORDER BY b.created_at DESC
    ''').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])
 
@app.route('/api/bills/<int:bid>', methods=['GET'])
def get_bill(bid):
    conn = get_db_joined()
    bill = conn.execute('''
        SELECT b.*, p.name as patient_name, p.age, p.gender, p.greeting, p.phone as patient_phone,
               p.blood_group, p.address,
               d.name as doctor_name, d.specialization, d.hospital
        FROM bills b
        LEFT JOIN patients p ON b.patient_id = p.id
        LEFT JOIN doctors d ON b.doctor_id = d.id
        WHERE b.id=?
    ''', (bid,)).fetchone()
    items = conn.execute("SELECT * FROM bill_items WHERE bill_id=? ORDER BY id", (bid,)).fetchall()
    settings = conn.execute("SELECT * FROM lab_settings LIMIT 1").fetchone()
    conn.close()
    if not bill:
        return jsonify({'error': 'Not found'}), 404
    return jsonify({
        'bill': dict(bill),
        'items': [dict(i) for i in items],
        'settings': dict(settings) if settings else {}
    })
 
 
@app.route('/api/bills', methods=['POST'])
def create_bill():
    data = request.get_json()
    conn = get_db_joined()
    # Resolve doctor_id from name if provided
    doctor_id = data.get('doctor_id')
    if not doctor_id and data.get('doctor_name') and data.get('doctor_name') != '—':
        row = conn.execute("SELECT id FROM doctors WHERE name=?", (data['doctor_name'],)).fetchone()
        if row:
            doctor_id = row['id']
    cur = conn.execute(
        "INSERT INTO bills (patient_id, doctor_id, bill_date, bill_type, cheque_ref, total_amount, status) VALUES (?,?,?,?,?,?,?)",
        (
            data.get('patient_id'),
            doctor_id,
            data.get('bill_date'),
            data.get('bill_type', 'Cash'),
            data.get('cheque_ref', ''),
            data.get('total_amount', 0),
            'saved'
        )
    )
    bill_id = cur.lastrowid
    for item in data.get('items', []):
        conn.execute(
            "INSERT INTO bill_items (bill_id, stage1_name, stage2_name, stage3_name, test_name, rate) VALUES (?,?,?,?,?,?)",
            (bill_id, item.get('stage1_name',''), item.get('stage2_name',''), item.get('stage3_name',''), item.get('test_name',''), item.get('rate', 0))
        )
    conn.commit()
    conn.close()
    return jsonify({'id': bill_id})
 
@app.route('/api/bills/<int:bid>', methods=['DELETE'])
def delete_bill(bid):
    conn = get_db()
    conn.execute("DELETE FROM bill_items WHERE bill_id=?", (bid,))
    conn.execute("DELETE FROM bills WHERE id=?", (bid,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Deleted'})
 
@app.route('/api/test-rates', methods=['GET'])
def get_all_test_rates():
    conn = get_db()
    rows = conn.execute("SELECT stage1_name, stage2_name, stage3_name, rate FROM test_rates").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])
 
@app.route('/api/test-rate', methods=['GET'])
def get_test_rate():
    s1 = request.args.get('s1', '')
    s2 = request.args.get('s2', '')
    s3 = request.args.get('s3', '')
    conn = get_db()
    row = conn.execute("SELECT rate FROM test_rates WHERE stage1_name=? AND stage2_name=? AND stage3_name=?",
                       (s1, s2, s3)).fetchone()
    conn.close()
    return jsonify({'rate': row['rate'] if row else 0})
 
 
# ── BACKUP & RESTORE ──────────────────────────────────────────────────────────
 
@app.route('/api/backup/download', methods=['GET'])
def backup_download():
    if not os.path.exists(DB_PATH):
        return jsonify({'error': 'Database not found'}), 404
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(DB_PATH, as_attachment=True,
                     download_name=f'labsoft_backup_{timestamp}.db',
                     mimetype='application/octet-stream')
 
@app.route('/api/backup/excel', methods=['GET'])
def backup_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    conn = get_db_joined()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
 
    header_font  = Font(bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill('solid', fgColor='1A4A7A')
    alt_fill     = PatternFill('solid', fgColor='EFF6FF')
    center       = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left         = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin         = Side(style='thin', color='BFDBFE')
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
 
    def make_sheet(title, rows, headers):
        ws = wb.create_sheet(title=title)
        ws.row_dimensions[1].height = 22
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font    = header_font
            cell.fill    = header_fill
            cell.alignment = center
            cell.border  = border
        for ri, row in enumerate(rows, 2):
            fill = alt_fill if ri % 2 == 0 else None
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.alignment = left
                cell.border    = border
                if fill:
                    cell.fill  = fill
        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        return ws
 
    # Patients
    pts = conn.execute("SELECT id,name,age,gender,phone,email,address,blood_group,created_at FROM patients ORDER BY id").fetchall()
    make_sheet('Patients', [tuple(r) for r in pts],
               ['ID','Name','Age','Gender','Phone','Email','Address','Blood Group','Created At'])
 
    # Doctors
    docs = conn.execute("SELECT id,name,specialization,hospital,phone,email,created_at FROM doctors ORDER BY id").fetchall()
    make_sheet('Doctors', [tuple(r) for r in docs],
               ['ID','Name','Specialization','Hospital','Phone','Email','Created At'])
 
    # Reports
    rpts = conn.execute('''SELECT r.id, p.name, d.name, r.report_title, r.report_date,
                                  r.status, r.created_at
                           FROM reports r
                           LEFT JOIN patients p ON r.patient_id=p.id
                           LEFT JOIN doctors  d ON r.doctor_id=d.id
                           ORDER BY r.id''').fetchall()
    make_sheet('Reports', [tuple(r) for r in rpts],
               ['ID','Patient','Doctor','Title','Date','Status','Created At'])
 
    # Bills
    bills = conn.execute('''SELECT b.id, p.name, d.name, b.bill_date, b.bill_type,
                                   b.total_amount, b.status, b.created_at
                            FROM bills b
                            LEFT JOIN patients p ON b.patient_id=p.id
                            LEFT JOIN doctors  d ON b.doctor_id=d.id
                            ORDER BY b.id''').fetchall()
    make_sheet('Bills', [tuple(r) for r in bills],
               ['ID','Patient','Doctor','Date','Type','Total (₹)','Status','Created At'])
 
    # Bill Items
    items = conn.execute('''SELECT bi.bill_id, p.name, bi.stage1_name, bi.stage2_name,
                                   bi.stage3_name, bi.test_name, bi.rate
                            FROM bill_items bi
                            JOIN bills b ON bi.bill_id=b.id
                            LEFT JOIN patients p ON b.patient_id=p.id
                            ORDER BY bi.bill_id''').fetchall()
    make_sheet('Bill Items', [tuple(r) for r in items],
               ['Bill ID','Patient','Category','Test Group','Sub-Test','Test Name','Rate (₹)'])
 
    # Appointments
    appts = conn.execute('''SELECT a.id, p.name, d.name, a.appointment_date, a.appointment_time,
                                   a.test_names, a.status, a.notes, a.created_at
                            FROM appointments a
                            LEFT JOIN patients p ON a.patient_id=p.id
                            LEFT JOIN doctors  d ON a.doctor_id=d.id
                            ORDER BY a.id''').fetchall()
    make_sheet('Appointments', [tuple(r) for r in appts],
               ['ID','Patient','Doctor','Date','Time','Tests','Status','Notes','Created At'])
 
    # Test Catalog
    tests = conn.execute("SELECT test_name,category,unit,normal_min,normal_max,normal_text,amount FROM test_catalog ORDER BY category,test_name").fetchall()
    make_sheet('Test Catalog', [tuple(r) for r in tests],
               ['Test Name','Category','Unit','Normal Min','Normal Max','Normal Range','Amount (₹)'])
 
    conn.close()
 
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'labsoft_backup_{timestamp}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
 
 
@app.route('/api/backup/pdf', methods=['GET'])
def backup_pdf():
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
 
    conn = get_db_joined()
    settings = dict(conn.execute("SELECT * FROM lab_settings LIMIT 1").fetchone() or {})
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    out = io.BytesIO()
 
    doc = SimpleDocTemplate(out, pagesize=landscape(A4),
                            leftMargin=12*mm, rightMargin=12*mm,
                            topMargin=0*mm, bottomMargin=18*mm,
                            canvasmaker=add_lab_logo_canvas(doc))
    styles = getSampleStyleSheet()
    primary   = colors.HexColor('#1A4A7A')
    accent    = colors.HexColor('#0d9488')
    alt_row   = colors.HexColor('#EFF6FF')
    hdr_fill  = colors.HexColor('#1A4A7A')
    story     = []
 
    title_s = ParagraphStyle('T', fontSize=18, textColor=primary, fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=2)
    sub_s   = ParagraphStyle('S', fontSize=9,  textColor=colors.grey, alignment=TA_CENTER, spaceAfter=2)
    sec_s   = ParagraphStyle('H', fontSize=12, textColor=primary, fontName='Helvetica-Bold', spaceBefore=12, spaceAfter=4)
    foot_s  = ParagraphStyle('F', fontSize=7,  textColor=colors.grey, alignment=TA_CENTER)
 
    lab_name = settings.get('lab_name','LabSoft')
    story.append(Paragraph(lab_name, title_s))
    story.append(Paragraph(f"Full Data Backup Report — {datetime.now().strftime('%d %b %Y %I:%M %p')}", sub_s))
    story.append(HRFlowable(width='100%', thickness=2, color=accent, spaceAfter=10))
 
    def tbl(headers, data, col_widths=None):
        hdr = [Paragraph(f'<b>{h}</b>', ParagraphStyle('th', fontSize=8, textColor=colors.white, alignment=TA_CENTER)) for h in headers]
        rows = [hdr]
        for i, row in enumerate(data):
            rows.append([Paragraph(str(v or '—'), ParagraphStyle('td', fontSize=8, alignment=TA_LEFT)) for v in row])
        t = Table(rows, colWidths=col_widths, repeatRows=1)
        style = [
            ('BACKGROUND', (0,0), (-1,0), hdr_fill),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, alt_row]),
            ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#BFDBFE')),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('PADDING', (0,0), (-1,-1), 4),
        ]
        t.setStyle(TableStyle(style))
        return t
 
    W = landscape(A4)[0] - 24*mm  # usable width
 
    # ── Patients ──
    pts = conn.execute("SELECT id,name,age,gender,phone,blood_group,created_at FROM patients ORDER BY id").fetchall()
    story.append(Paragraph(f'👤 Patients ({len(pts)})', sec_s))
    if pts:
        story.append(tbl(['ID','Name','Age','Gender','Phone','Blood Group','Registered'],
                         [tuple(r) for r in pts],
                         [15*mm, 55*mm, 15*mm, 20*mm, 35*mm, 25*mm, 40*mm]))
    else:
        story.append(Paragraph('No records.', styles['Normal']))
 
    # ── Doctors ──
    docs = conn.execute("SELECT id,name,specialization,hospital,phone FROM doctors ORDER BY id").fetchall()
    story.append(Paragraph(f'✚ Doctors ({len(docs)})', sec_s))
    if docs:
        story.append(tbl(['ID','Name','Specialization','Hospital','Phone'],
                         [tuple(r) for r in docs],
                         [15*mm, 55*mm, 50*mm, 60*mm, 35*mm]))
    else:
        story.append(Paragraph('No records.', styles['Normal']))
 
    # ── Reports ──
    rpts = conn.execute('''SELECT r.id, p.name, d.name, r.report_title, r.report_date, r.status
                           FROM reports r
                           LEFT JOIN patients p ON r.patient_id=p.id
                           LEFT JOIN doctors  d ON r.doctor_id=d.id ORDER BY r.id''').fetchall()
    story.append(Paragraph(f'📋 Reports ({len(rpts)})', sec_s))
    if rpts:
        story.append(tbl(['ID','Patient','Doctor','Title','Date','Status'],
                         [tuple(r) for r in rpts],
                         [15*mm, 50*mm, 45*mm, 60*mm, 28*mm, 22*mm]))
    else:
        story.append(Paragraph('No records.', styles['Normal']))
 
    # ── Bills ──
    bills = conn.execute('''SELECT b.id, p.name, b.bill_date, b.bill_type, b.total_amount, b.status
                            FROM bills b LEFT JOIN patients p ON b.patient_id=p.id ORDER BY b.id''').fetchall()
    story.append(Paragraph(f'🧾 Bills ({len(bills)})', sec_s))
    if bills:
        story.append(tbl(['ID','Patient','Date','Type','Total (₹)','Status'],
                         [tuple(r) for r in bills],
                         [15*mm, 60*mm, 28*mm, 25*mm, 30*mm, 25*mm]))
    else:
        story.append(Paragraph('No records.', styles['Normal']))
 
    # ── Appointments ──
    appts = conn.execute('''SELECT a.id, p.name, d.name, a.appointment_date, a.appointment_time, a.status
                            FROM appointments a
                            LEFT JOIN patients p ON a.patient_id=p.id
                            LEFT JOIN doctors  d ON a.doctor_id=d.id ORDER BY a.id''').fetchall()
    story.append(Paragraph(f'📅 Appointments ({len(appts)})', sec_s))
    if appts:
        story.append(tbl(['ID','Patient','Doctor','Date','Time','Status'],
                         [tuple(r) for r in appts],
                         [15*mm, 55*mm, 50*mm, 28*mm, 22*mm, 25*mm]))
    else:
        story.append(Paragraph('No records.', styles['Normal']))
 
    conn.close()
 
    foot_text = f"LabSoft Backup PDF · {datetime.now().strftime('%d %b %Y %I:%M %p')} · {lab_name}"
    def draw_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(colors.grey)
        canvas.line(12*mm, 12*mm, landscape(A4)[0]-12*mm, 12*mm)
        canvas.drawCentredString(landscape(A4)[0]/2, 7*mm, foot_text)
        canvas.restoreState()
 
    doc.build(story, onFirstPage=draw_footer, onLaterPages=draw_footer)
    out.seek(0)
    return send_file(out, as_attachment=True,
                     download_name=f'labsoft_backup_{timestamp}.pdf',
                     mimetype='application/pdf')
 
@app.route('/api/backup/restore', methods=['POST'])
def backup_restore():
    file = request.files.get('db_file')
    if not file or not file.filename.endswith('.db'):
        return jsonify({'error': 'Please upload a valid .db file'}), 400
    # Save a safety copy of current DB first
    if os.path.exists(DB_PATH):
        safety = DB_PATH + '.bak'
        import shutil
        shutil.copy2(DB_PATH, safety)
    try:
        file.save(DB_PATH)
        # Quick sanity check — open and query
        conn = sqlite3.connect(DB_PATH)
        conn.execute("SELECT COUNT(*) FROM patients")
        conn.close()
        return jsonify({'message': 'Database restored successfully'})
    except Exception as e:
        # Rollback to safety copy
        if os.path.exists(DB_PATH + '.bak'):
            import shutil
            shutil.copy2(DB_PATH + '.bak', DB_PATH)
        return jsonify({'error': f'Restore failed: {str(e)}'}), 500
 
@app.route('/api/backup/stats', methods=['GET'])
def backup_stats():
    conn = get_db_joined()
    stats = {
        'patients': conn.execute("SELECT COUNT(*) FROM patients").fetchone()[0],
        'reports':  conn.execute("SELECT COUNT(*) FROM reports").fetchone()[0],
        'doctors':  conn.execute("SELECT COUNT(*) FROM doctors").fetchone()[0],
        'bills':    conn.execute("SELECT COUNT(*) FROM bills").fetchone()[0],
    }
    conn.close()
    return jsonify(stats)
 
 
 
# ── OTHER LABS ────────────────────────────────────────────────────────────────
 
@app.route('/api/other-labs', methods=['GET'])
def get_other_labs():
    conn = get_db()
    rows = conn.execute("SELECT * FROM other_labs ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])
 
@app.route('/api/other-labs', methods=['POST'])
def add_other_lab():
    data = request.json
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Lab name required'}), 400
    conn = get_db()
    try:
        c = conn.cursor()
        c.execute("INSERT INTO other_labs (name) VALUES (?)", (name,))
        lid = c.lastrowid
        conn.commit()
        conn.close()
        return jsonify({'id': lid, 'message': 'Lab added'})
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 400
 
@app.route('/api/other-labs/<int:lid>', methods=['DELETE'])
def delete_other_lab(lid):
    conn = get_db()
    conn.execute("DELETE FROM other_lab_ranges WHERE lab_id=?", (lid,))
    conn.execute("DELETE FROM other_labs WHERE id=?", (lid,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Deleted'})
 
# ── OTHER LAB RANGES ──────────────────────────────────────────────────────────
 
@app.route('/api/other-lab-ranges', methods=['GET'])
def get_other_lab_ranges():
    lab_id = request.args.get('lab_id', '')
    q = request.args.get('q', '')
    conn = get_db()
    sql = "SELECT * FROM other_lab_ranges WHERE lab_id=?"
    params = [lab_id]
    if q:
        sql += " AND test_name LIKE ?"
        params.append(f'%{q}%')
    sql += " ORDER BY category, test_name"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])
 
@app.route('/api/other-lab-ranges', methods=['POST'])
def add_other_lab_range():
    data = request.json
    conn = get_db()
    c = conn.cursor()
    c.execute('''INSERT INTO other_lab_ranges
        (lab_id, test_name, category, unit,
         normal_min_m, normal_max_m, normal_text_m,
         normal_min_f, normal_max_f, normal_text_f,
         normal_min_c, normal_max_c, normal_text_c,
         description, interpretation, amount)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (data['lab_id'], data['test_name'], data.get('category'), data.get('unit'),
         data.get('normal_min_m') or None, data.get('normal_max_m') or None, data.get('normal_text_m'),
         data.get('normal_min_f') or None, data.get('normal_max_f') or None, data.get('normal_text_f'),
         data.get('normal_min_c') or None, data.get('normal_max_c') or None, data.get('normal_text_c'),
         data.get('description'), data.get('interpretation'), data.get('amount') or 0))
    rid = c.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'id': rid, 'message': 'Range added'})
 
@app.route('/api/other-lab-ranges/<int:rid>', methods=['PUT'])
def update_other_lab_range(rid):
    data = request.json
    conn = get_db()
    conn.execute('''UPDATE other_lab_ranges SET
        test_name=?, category=?, unit=?,
        normal_min_m=?, normal_max_m=?, normal_text_m=?,
        normal_min_f=?, normal_max_f=?, normal_text_f=?,
        normal_min_c=?, normal_max_c=?, normal_text_c=?,
        description=?, interpretation=?, amount=?
        WHERE id=?''',
        (data['test_name'], data.get('category'), data.get('unit'),
         data.get('normal_min_m') or None, data.get('normal_max_m') or None, data.get('normal_text_m'),
         data.get('normal_min_f') or None, data.get('normal_max_f') or None, data.get('normal_text_f'),
         data.get('normal_min_c') or None, data.get('normal_max_c') or None, data.get('normal_text_c'),
         data.get('description'), data.get('interpretation'), data.get('amount') or 0, rid))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Updated'})
 
@app.route('/api/other-lab-ranges/<int:rid>', methods=['DELETE'])
def delete_other_lab_range(rid):
    conn = get_db()
    conn.execute("DELETE FROM other_lab_ranges WHERE id=?", (rid,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Deleted'})
 
 
import webbrowser
from threading import Timer
 
# ── LICENSE ROUTES ────────────────────────────────────────────────────────────
 
@app.route('/api/license/status', methods=['GET'])
def license_status():
    status = check_license_status()
    # Replace raw machine_id with encrypted system_code for the client UI
    if status.get('machine_id'):
        status['system_code'] = get_system_code(status['machine_id'])
    return jsonify(status)
 
@app.route('/api/license/activate', methods=['POST'])
def license_activate():
    data = request.get_json(silent=True) or {}
    license_key = str(data.get('license_key', '')).strip()
    if not license_key:
        return jsonify({'success': False, 'message': 'License key is required.'}), 400
    machine_id = get_machine_id()
    success, message = activate_license(machine_id, license_key)
    return jsonify({'success': success, 'message': message})

@app.route('/api/license/machine-id', methods=['GET'])
def license_machine_id():
    return jsonify({'machine_id': get_machine_id()})
 
@app.route('/api/session/count', methods=['GET'])
def get_session_count():
    return jsonify({'count': _read_session_count()})
 
# Inside your main block
if __name__ == "__main__":
    port = 5000
    # Increment session counter if license is valid
    _startup_status = check_license_status()
    if _startup_status.get('valid'):
        increment_session_count()
    # Open browser after 1 second
    Timer(1, lambda: webbrowser.open(f"http://127.0.0.1:{port}")).start()
    app.run(debug=False, port=port)
